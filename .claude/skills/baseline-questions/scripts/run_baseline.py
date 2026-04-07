"""Run EAGLE baseline questions against a live server and save results to Excel.

Usage:
    python run_baseline.py --version v5 [--server URL] [--xlsx PATH] [--tenant ID]
"""
import argparse
import asyncio
import json
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")


def load_questions_from_excel(xlsx_path: str) -> dict:
    """Read questions from column D of the 'Baseline questions' sheet.

    Returns {row_number: question_text} for every non-empty cell in column D
    starting from row 2.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Baseline questions"]
    questions = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=4).value
        if val and str(val).strip():
            questions[row] = str(val).strip()
    wb.close()
    return questions


def find_next_column(ws) -> int:
    """Find the first empty column in row 1."""
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        col += 1
    return col


async def run_question(
    client: httpx.AsyncClient,
    base_url: str,
    tenant: str,
    row: int,
    question: str,
) -> dict:
    """Send a question to the EAGLE server and return the response."""
    session_id = str(uuid.uuid4())
    q_num = row - 1

    print(f"\n{'='*80}")
    print(f"Q{q_num} (row {row}): {question[:80]}...")
    print(f"Session: {session_id}")
    print(f"{'='*80}")

    start = time.time()
    try:
        resp = await client.post(
            f"{base_url}/api/chat",
            json={"message": question, "session_id": session_id},
            headers={
                "X-User-Id": "baseline-eval",
                "X-Tenant-Id": tenant,
                "X-User-Email": "baseline@eval.test",
                "X-User-Tier": "advanced",
            },
            timeout=300.0,
        )
        elapsed = time.time() - start
        data = resp.json()

        response_text = data.get("response", "")
        tools = data.get("tools_called", [])
        usage = data.get("usage", {})
        model = data.get("model", "unknown")

        print(f"\nCompleted in {elapsed:.1f}s | Model: {model}")
        print(f"Tools: {tools}")
        print(f"Tokens: in={usage.get('input_tokens', 0):,} out={usage.get('output_tokens', 0):,}")
        print(f"Response length: {len(response_text):,} chars")
        print(f"\nFull response:\n{response_text}")

        return {
            "row": row,
            "q_num": q_num,
            "response": response_text,
            "tools": tools,
            "usage": usage,
            "model": model,
            "elapsed_s": round(elapsed, 1),
            "session_id": session_id,
            "status": "ok",
        }
    except Exception as e:
        elapsed = time.time() - start
        print(f"\nERROR after {elapsed:.1f}s: {e}")
        return {
            "row": row,
            "q_num": q_num,
            "response": f"ERROR: {e}",
            "tools": [],
            "usage": {},
            "model": "error",
            "elapsed_s": round(elapsed, 1),
            "session_id": session_id,
            "status": "error",
        }


def _parse_question_spec(spec: str) -> list[int]:
    """Parse a question spec like '1,3,5' or '7-10' or '2,8-12' into Q numbers."""
    nums: list[int] = []
    for part in spec.split(","):
        part = part.strip().lstrip("qQ")
        if "-" in part:
            lo, hi = part.split("-", 1)
            nums.extend(range(int(lo), int(hi) + 1))
        else:
            nums.append(int(part))
    return sorted(set(nums))


async def main():
    parser = argparse.ArgumentParser(description="Run EAGLE baseline evaluation")
    parser.add_argument("--version", required=True, help="Version label (e.g., v5)")
    parser.add_argument("--server", default="http://localhost:8000", help="Server URL")
    parser.add_argument("--xlsx", default=None, help="Excel workbook path")
    parser.add_argument("--tenant", default="dev-tenant", help="Tenant ID")
    parser.add_argument(
        "--questions", default=None,
        help="Run specific questions only. Comma-separated Q numbers (e.g., 1,3,5) "
             "or a range (e.g., 7-10). Omit to run all.",
    )
    args = parser.parse_args()

    # Default xlsx path: repo root / Use Case List.xlsx
    if args.xlsx:
        xlsx_path = args.xlsx
    else:
        # Walk up from this script to find repo root
        repo_root = Path(__file__).resolve().parent.parent.parent.parent.parent
        xlsx_path = str(repo_root / "Use Case List.xlsx")

    today = datetime.now().strftime("%Y-%m-%d")
    version = args.version.upper() if not args.version[0].isupper() else args.version

    # ── Load questions from Excel ──
    ALL_QUESTIONS = load_questions_from_excel(xlsx_path)
    if not ALL_QUESTIONS:
        print("ERROR: No questions found in column D of 'Baseline questions' sheet")
        sys.exit(1)

    # ── Filter to specific questions if --questions flag provided ──
    if args.questions:
        requested = _parse_question_spec(args.questions)
        # Map Q numbers to Excel rows (Q1=row2, Q2=row3, etc.)
        all_by_qnum = {row - 1: row for row in ALL_QUESTIONS}
        QUESTIONS = {}
        for qn in requested:
            if qn in all_by_qnum:
                row = all_by_qnum[qn]
                QUESTIONS[row] = ALL_QUESTIONS[row]
            else:
                print(f"WARNING: Q{qn} not found in Excel (available: Q{min(all_by_qnum)}-Q{max(all_by_qnum)})")
        if not QUESTIONS:
            print("ERROR: None of the requested questions exist")
            sys.exit(1)
        print(f"Running {len(QUESTIONS)} of {len(ALL_QUESTIONS)} questions: "
              f"Q{', Q'.join(str(r-1) for r in sorted(QUESTIONS))}")
    else:
        QUESTIONS = ALL_QUESTIONS

    print(f"EAGLE {version} Baseline Evaluation")
    print(f"Server: {args.server}")
    print(f"Excel:  {xlsx_path}")
    print(f"Tenant: {args.tenant}")
    print(f"Date:   {today}")
    print(f"Questions: {len(QUESTIONS)}")

    # ── Preflight: check server ──
    async with httpx.AsyncClient() as client:
        try:
            r = await client.get(f"{args.server}/api/health", timeout=5)
            health = r.json()
            print(f"Server: {health.get('service', '?')} {health.get('version', '?')} - OK")
        except Exception as e:
            print(f"\nERROR: Server not reachable at {args.server}")
            print(f"  {e}")
            print(f"\nStart the server first:")
            print(f"  cd server && uvicorn app.main:app --reload --port 8000")
            sys.exit(1)

    # ── Open workbook and read RO reference responses from column E ──
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Baseline questions"]

    ro_responses = {}
    for row in sorted(QUESTIONS.keys()):
        ro_val = ws.cell(row=row, column=5).value
        ro_responses[row] = str(ro_val) if ro_val else ""
    ro_count = sum(1 for v in ro_responses.values() if v)
    print(f"RO reference responses loaded: {ro_count}/{len(QUESTIONS)} from column E")

    # ── Run questions sequentially ──
    results = {}
    async with httpx.AsyncClient() as client:
        for row in sorted(QUESTIONS.keys()):
            result = await run_question(client, args.server, args.tenant, row, QUESTIONS[row])
            result["ro_response"] = ro_responses.get(row, "")
            result["question"] = QUESTIONS[row]
            results[row] = result

    # ── Save raw JSON ──
    json_dir = Path(__file__).resolve().parent.parent.parent.parent.parent / "scripts"
    json_dir.mkdir(exist_ok=True)
    json_path = json_dir / f"baseline_{args.version}_results.json"
    serializable = {str(row): r for row, r in results.items()}
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)
    print(f"\nRaw results saved to {json_path}")

    # ── Write to Excel ──
    print(f"\nWriting to {xlsx_path}...")
    col = find_next_column(ws)
    wrap = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Header
    header_text = f"EAGLE {version} Response ({today})"
    cell = ws.cell(row=1, column=col, value=header_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

    # Responses
    for row, r in results.items():
        ws.cell(row=row, column=col, value=r["response"]).alignment = wrap

    # Set column width
    col_letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 100

    wb.save(xlsx_path)
    print(f"Done! EAGLE {version} responses written to column {col_letter} (col {col})")

    # ── Summary ──
    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    print(f"{'Q#':<5} {'Status':<8} {'Time':>6} {'EAGLE':>8} {'RO Ref':>8}  Tools")
    print("-" * 80)
    for row in sorted(results.keys()):
        r = results[row]
        status = "OK" if r["status"] == "ok" else "ERROR"
        ro_len = len(r.get("ro_response", ""))
        print(
            f"Q{r['q_num']:<4} {status:<8} {r['elapsed_s']:>5.1f}s {len(r['response']):>7,} {ro_len:>7,}  {r['tools']}"
        )

    total_time = sum(r["elapsed_s"] for r in results.values())
    total_chars = sum(len(r["response"]) for r in results.values())
    total_ro = sum(len(r.get("ro_response", "")) for r in results.values())
    errors = sum(1 for r in results.values() if r["status"] == "error")
    print(f"\nTotal: {total_time:.0f}s | EAGLE {total_chars:,} chars | RO {total_ro:,} chars | {errors} errors")
    print(f"Response column: {col_letter} (col {col})")
    print(f"JSON: {json_path}")
    print(f"\nRO reference responses (column E) included in JSON for judging comparison.")

    # ── Tool usage analysis ──
    print(f"\n{'='*80}")
    print("TOOL USAGE ANALYSIS")
    print(f"{'='*80}")

    fetch_questions = []
    no_fetch_questions = []
    for row in sorted(results.keys()):
        r = results[row]
        tools = r["tools"]
        has_fetch = "knowledge_fetch" in tools
        has_search = "knowledge_search" in tools or "search_far" in tools

        if has_search and has_fetch:
            fetch_questions.append(r["q_num"])
        elif has_search and not has_fetch:
            no_fetch_questions.append(r["q_num"])

    if fetch_questions:
        print(f"  knowledge_fetch called after search: Q{', Q'.join(str(q) for q in fetch_questions)}")
    if no_fetch_questions:
        print(f"  WARNING: search without fetch: Q{', Q'.join(str(q) for q in no_fetch_questions)}")
        print(f"  (cascade enforcement may not be working for these questions)")
    if not fetch_questions and not no_fetch_questions:
        print("  No KB search tools called (Q1=compliance_matrix, Q6=no tools)")


if __name__ == "__main__":
    asyncio.run(main())
