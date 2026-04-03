"""Run baseline questions against local EAGLE server and save responses to Excel."""
import asyncio
import json
import sys
import time
import uuid

import httpx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

BASE_URL = "http://localhost:8000"
XLSX = "C:/Users/blackga/Desktop/eagle/sm_eagle/Use Case List.xlsx"

QUESTIONS = {
    2: "What are the simplified acquisition threshold and micro-purchase threshold under FAC 2025-06?",
    3: "What did GAO hold in B-302358 regarding IDIQ minimum obligation requirements?",
    4: "How does the severable vs. non-severable distinction affect which fiscal year's appropriation must fund a contract?",
    5: "What are the fair opportunity exceptions that allow a single-award task order under FAR 16.505?",
    6: (
        "An offeror eliminated from a SBIR competition simultaneously requests a debriefing "
        "and files a protest on Day 8 after receiving the elimination notice. The debriefing "
        "hasn't been provided yet. Walk through the correct procedural sequence, applicable "
        "protest timeliness rules, and how the choice between pre-award and post-award "
        "debriefing affects the protest stay and timeline."
    ),
    7: (
        "were coming back to the original questions. how to start that discussion - we told "
        "eagle we wanted to buy cloud services and it shot out a quick SOW with only a few "
        "questions. not bad. but we think about what if i went to price it and it was too "
        "expensive and I had to go backwards. we have those multiple points of entry and "
        "balance before we can get to the end of the document workflow. we used to come in "
        "with figuring out the background and objectives etc. that was a good approach but "
        "it did sometimes get distracted and chat theory forever. hard to tune."
    ),
}


async def run_question(client: httpx.AsyncClient, row: int, question: str) -> dict:
    """Send a question to the EAGLE server and return the response."""
    session_id = str(uuid.uuid4())
    q_num = row - 1

    print(f"\n{'='*80}")
    print(f"Q{q_num} (row {row}): {question[:80]}...")
    print(f"Session: {session_id}")
    print(f"{'='*80}")

    start = time.time()
    try:
        resp = await client.post(
            f"{BASE_URL}/api/chat",
            json={
                "message": question,
                "session_id": session_id,
            },
            headers={
                "X-User-Id": "baseline-eval",
                "X-Tenant-Id": "dev-tenant",
                "X-User-Email": "baseline@eval.test",
                "X-User-Tier": "advanced",
            },
            timeout=300.0,  # 5 min per question
        )
        elapsed = time.time() - start
        data = resp.json()

        response_text = data.get("response", "")
        tools = data.get("tools_called", [])
        usage = data.get("usage", {})
        model = data.get("model", "unknown")

        print(f"\nCompleted in {elapsed:.1f}s | Model: {model}")
        print(f"Tools: {tools}")
        print(f"Tokens: in={usage.get('input_tokens', 0):,} out={usage.get('output_tokens', 0):,}")
        print(f"Response length: {len(response_text):,} chars")
        print(f"\nFirst 500 chars:\n{response_text[:500]}")

        return {
            "row": row,
            "response": response_text,
            "tools": tools,
            "usage": usage,
            "model": model,
            "elapsed_s": round(elapsed, 1),
            "session_id": session_id,
            "status": "ok",
        }

    except Exception as e:
        elapsed = time.time() - start
        print(f"\nERROR after {elapsed:.1f}s: {e}")
        return {
            "row": row,
            "response": f"ERROR: {e}",
            "tools": [],
            "usage": {},
            "model": "error",
            "elapsed_s": round(elapsed, 1),
            "session_id": session_id,
            "status": "error",
        }


async def main():
    print("EAGLE v5 Baseline Evaluation")
    print(f"Server: {BASE_URL}")
    print(f"Questions: {len(QUESTIONS)}")
    print()

    # Run questions sequentially (each needs full model attention)
    results = {}
    async with httpx.AsyncClient() as client:
        for row in sorted(QUESTIONS.keys()):
            result = await run_question(client, row, QUESTIONS[row])
            results[row] = result

    # Save raw JSON responses for reference
    json_path = "C:/Users/blackga/Desktop/eagle/sm_eagle/scripts/baseline_v5_results.json"
    serializable = {}
    for row, r in results.items():
        serializable[str(row)] = r
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)
    print(f"\nRaw results saved to {json_path}")

    # Write to Excel
    print(f"\nWriting to {XLSX}...")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Baseline questions"]

    col = 16  # Column P — EAGLE v5 Response
    wrap = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Header
    cell = ws.cell(row=1, column=col, value="EAGLE v5 Response (2026-04-02)")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

    # Responses
    for row, r in results.items():
        ws.cell(row=row, column=col, value=r["response"]).alignment = wrap

    ws.column_dimensions["P"].width = 100
    wb.save(XLSX)
    print("Done! EAGLE v5 responses written to column P.")

    # Summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    for row in sorted(results.keys()):
        r = results[row]
        q_num = row - 1
        status = "OK" if r["status"] == "ok" else "ERROR"
        print(
            f"Q{q_num}: {status} | {r['elapsed_s']}s | "
            f"{len(r['response']):,} chars | tools: {r['tools']}"
        )


if __name__ == "__main__":
    asyncio.run(main())
