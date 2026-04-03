"""Score v7 baseline responses and write to Excel."""
import sys, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.load_workbook("Use Case List.xlsx")
ws = wb["Baseline questions"]

SCORE_START = 31
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
wrap = Alignment(wrap_text=True, vertical="top")

headers = [
    "EAGLE V7 Accuracy (0-5)",
    "EAGLE V7 Completeness (0-5)",
    "EAGLE V7 Sources (0-5)",
    "EAGLE V7 Actionability (0-5)",
    "EAGLE V7 Total (0-20)",
    "v7 vs v6 Comparative Judgment",
]
for i, h in enumerate(headers):
    cell = ws.cell(row=1, column=SCORE_START + i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

# Score data: row -> (acc, comp, src, act, judgment)
scores = {
    2: (5, 5, 4, 5,
        "v7 = v6 (no change)\n\n"
        "Both produce correct threshold tables: $15K MPT, $350K SAT, construction/SCA/contingency "
        "exceptions, commercial items $7.5M/$15M. v7 slightly more concise (711 vs 1139 chars). "
        "v6 included bonus thresholds ($25K SAM.gov synopsis, $150K VETS-4212) not strictly asked for.\n\n"
        "Tools: v6=query_compliance_matrix | v7=query_compliance_matrix\n"
        "Cascade: Q1 uses compliance matrix directly, no KB cascade needed. No change.\n"
        "Source gaps: None -- matrix is authoritative for thresholds.\n\n"
        "Verdict: Equivalent. Both answer the specific question completely."),

    3: (5, 5, 4, 5,
        "v7 = v6 (comparable, minor tool variation)\n\n"
        "Both provide comprehensive B-302358 analysis. v7 structures 4 holdings (minimum obligation, "
        "task order timing, IDIQ vs multiyear exclusivity, substance-over-label). v6 provides more case "
        "background (IBM/ACE $25M min, $5B max). Both cite 31 USC 1501(a)(1), Crown Laundry.\n\n"
        "Tools: v6=knowledge_search,web_search,web_fetch | v7=knowledge_search,legal_counsel\n"
        "Cascade: v6 web-fetched actual decision; v7 used legal_counsel subagent instead. Both valid "
        "approaches. Minor sources deduction for not fetching original text.\n\n"
        "Verdict: Comparable. v7 loses 1pt on sources (no web fetch of GAO decision) but gains "
        "structured legal analysis from subagent."),

    4: (5, 5, 5, 5,
        "v7 = v6 (comparable)\n\n"
        "Both provide thorough severable/non-severable analysis with correct core rules, funding "
        "examples, GAO citations. v7 uses the research tool (composite: auto-search + auto-fetch). "
        "v7 adds a 'Gray Zone: IT Contracts' table -- highly practical for NIH. v6 includes the "
        "12-month crossing exception; v7 mentions the short bridge exception.\n\n"
        "Tools: v6=knowledge_search,knowledge_fetch | v7=research\n"
        "Cascade: CASCADE WORKING -- v7 research tool auto-fetches top 4 KB documents.\n"
        "Source gaps: None -- both cite GAO Red Book Ch 5, B-317636, B-317139, FAR 32.703-2, 37.106.\n\n"
        "Verdict: Equivalent. v7's IT gray zone table is a strong practical addition. Both 20/20."),

    5: (5, 5, 5, 5,
        "v7 = v6 (comparable)\n\n"
        "Both provide complete FAR 16.507-6(b) analysis with all 6 exceptions, J&A approval thresholds, "
        "and procedural requirements. Nearly identical content and structure. v7 adds HHS Class Deviation "
        "2026-01 citation and full KB file path.\n\n"
        "Tools: v6=search_far,knowledge_fetch | v7=search_far,knowledge_fetch\n"
        "Cascade: CASCADE WORKING -- both fetch full FAR Part 16 text after search. Identical pattern.\n"
        "Source gaps: Closed -- full FAR Part 16 text read before answering in both versions.\n\n"
        "Verdict: Equivalent. Both 20/20. Identical cascade behavior."),

    6: (5, 5, 5, 5,
        "v7 = v6 (comparable)\n\n"
        "Both produce exceptional SBIR protest analysis (13-15K chars). v7 uses 5 tools including "
        "web_search/web_fetch to verify primary sources (4 CFR Part 21, 31 USC 3553, FAR 15.505/15.506). "
        "v7 adds SBIR-specific OTA jurisdiction note and consolidated timeline with specific day counts. "
        "v6 structured as Parts I-IV; v7 has 11 sections with more granular analysis.\n\n"
        "Tools: v6=knowledge_search,search_far,knowledge_fetch | v7=search_far,research,knowledge_fetch,web_search,web_fetch\n"
        "Cascade: CASCADE WORKING -- both go deep. v7 adds web verification of primary statutory sources.\n"
        "Source gaps: None -- both cite 4 CFR Part 21, 31 USC 3553, FAR 15.505/15.506, GAO-18-510SP.\n\n"
        "Verdict: Equivalent. Both 20/20. v7 adds web-verified primary sources and SBIR jurisdiction nuance."),

    7: (5, 5, 3, 5,
        "v7 = v6 (comparable)\n\n"
        "Both correctly diagnose the SOW-first rework problem and propose cost-validation gating. "
        "v7 offers a concrete 3-stage gating model (Framing -> Strategy Check -> Document Generation) "
        "with the insight that IGCE is the anchor document. v6 offers a 3-failure-modes diagnostic.\n\n"
        "Tools: v6=get_intake_status | v7=none\n"
        "Cascade: N/A -- design discussion. v7 correctly uses no tools. v6 called get_intake_status "
        "(unnecessary).\n"
        "Source gaps: N/A -- design discussion, no regulatory sources expected.\n\n"
        "Verdict: Equivalent. Both 18/20. v7 staged gate model slightly more concrete."),
}

for row, (acc, comp, src, act, judgment) in scores.items():
    total = acc + comp + src + act
    ws.cell(row=row, column=SCORE_START, value=acc).alignment = wrap
    ws.cell(row=row, column=SCORE_START+1, value=comp).alignment = wrap
    ws.cell(row=row, column=SCORE_START+2, value=src).alignment = wrap
    ws.cell(row=row, column=SCORE_START+3, value=act).alignment = wrap
    ws.cell(row=row, column=SCORE_START+4, value=total).alignment = wrap
    ws.cell(row=row, column=SCORE_START+5, value=judgment).alignment = wrap

for i in range(5):
    letter = openpyxl.utils.get_column_letter(SCORE_START + i)
    ws.column_dimensions[letter].width = 14
judgment_letter = openpyxl.utils.get_column_letter(SCORE_START + 5)
ws.column_dimensions[judgment_letter].width = 100

wb.save("Use Case List.xlsx")
print("Scores and judgments written to Excel!")

# ── Summary ──
print()
print("=" * 90)
print("V7 BASELINE EVALUATION SUMMARY")
print("=" * 90)
print(f"{'Q#':<5} {'Acc':>4} {'Comp':>5} {'Src':>4} {'Act':>4} {'Total':>6} {'v6 Tot':>7} {'Delta':>6}  Verdict")
print("-" * 90)

v6_scores = {2: 19, 3: 20, 4: 20, 5: 20, 6: 20, 7: 18}
v7_totals = []
for row in sorted(scores.keys()):
    acc, comp, src, act, _ = scores[row]
    total = acc + comp + src + act
    v7_totals.append(total)
    v6t = v6_scores[row]
    delta = total - v6t
    delta_s = f"+{delta}" if delta > 0 else str(delta)
    verdict = "v7 > v6" if delta > 0 else ("v7 < v6" if delta < 0 else "v7 = v6")
    print(f"Q{row-1:<4} {acc:>4} {comp:>5} {src:>4} {act:>4} {total:>5}/20 {v6t:>5}/20 {delta_s:>6}  {verdict}")

v7_avg = sum(v7_totals) / len(v7_totals)
v6_avg = sum(v6_scores.values()) / len(v6_scores)
print("-" * 90)
print(f"{'AVG':<5} {'':>4} {'':>5} {'':>4} {'':>4} {v7_avg:>5.1f}/20 {v6_avg:>5.1f}/20 {v7_avg-v6_avg:>+5.1f}")
print()

wins = sum(1 for r in scores if sum(scores[r][:4]) > v6_scores[r])
ties = sum(1 for r in scores if sum(scores[r][:4]) == v6_scores[r])
losses = sum(1 for r in scores if sum(scores[r][:4]) < v6_scores[r])
print(f"v7 wins: {wins}/6 | Ties: {ties}/6 | v6 wins: {losses}/6")

print()
print("KB CASCADE ENFORCEMENT IMPACT")
print("=" * 90)
v7_tools = {
    1: ["query_compliance_matrix"],
    2: ["knowledge_search", "legal_counsel"],
    3: ["research"],
    4: ["search_far", "knowledge_fetch"],
    5: ["search_far", "research", "knowledge_fetch", "web_search", "web_fetch"],
    6: [],
}
v6_tools = {
    1: ["query_compliance_matrix"],
    2: ["knowledge_search", "web_search", "web_fetch"],
    3: ["knowledge_search", "knowledge_fetch"],
    4: ["search_far", "knowledge_fetch"],
    5: ["knowledge_search", "search_far", "knowledge_fetch"],
    6: ["get_intake_status"],
}
for q in range(1, 7):
    v6t = v6_tools[q]
    v7t = v7_tools[q]
    changed = "CHANGED" if v6t != v7t else "same"
    print(f"  Q{q}: v6={v6t} | v7={v7t}  [{changed}]")

print()
print("Tool pattern changes:")
print("  Q2: v6 used web_search+web_fetch; v7 used legal_counsel subagent (different research path)")
print("  Q3: v6 used explicit search+fetch; v7 used research tool (composite, auto-fetches)")
print("  Q5: v7 added web_search+web_fetch for primary source verification (improvement)")
print("  Q6: v7 correctly uses NO tools (design discussion); v6 called get_intake_status (unnecessary)")
print()
print("Checklist isolation: v7 uses research tool for Q3 which includes isolated checklist fetch.")
print("No checklists leaked into general KB results (verified in prior testing).")
