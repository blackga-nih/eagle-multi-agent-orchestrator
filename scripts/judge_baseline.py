"""Judge EAGLE v4 baseline responses against prior system responses."""
import openpyxl
import sys
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

XLSX = "C:/Users/blackga/Desktop/eagle/sm_eagle/Use Case List.xlsx"
wb = openpyxl.load_workbook(XLSX)
ws = wb["Baseline questions"]

# New columns
COLS = {
    9: "EAGLE v4 Accuracy (0-5)",
    10: "EAGLE v4 Completeness (0-5)",
    11: "EAGLE v4 Sources (0-5)",
    12: "EAGLE v4 Actionability (0-5)",
    13: "EAGLE v4 Total (0-20)",
    14: "Comparative Judgment vs Baselines",
}

wrap = Alignment(wrap_text=True, vertical="top")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, label in COLS.items():
    cell = ws.cell(row=1, column=col_num, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

ratings = {
    2: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 4,
        "actionability": 4,
        "judgment": (
            "WINNER: EAGLE v4 (tie with GB MVP1)\n\n"
            "All three systems correctly identify MPT=$15K and SAT=$350K under FAC 2025-06.\n\n"
            "Research Optimizer (Col 5): Correct core answer with practical notes on above-MPT and above-SAT "
            "requirements. Concise but lacks exception thresholds. Sources KB doc but no FAR section numbers.\n\n"
            "GB MVP1 (Col 7): Adds important exceptions (construction $2K, SCA $2.5K, contingency $1.5M, "
            "commercial $7.5M). Cites FAR 2.101. Good practical depth.\n\n"
            "EAGLE v4 (Col 8): Formatted table with FAR references. Includes construction, SCA, contingency "
            "($25K - note: different from MVP1 $1.5M which is the SAT contingency exception, not MPT). "
            "Adds commercial COTS ceiling ($15M). Inflation adjustment warning is useful.\n\n"
            "EDGE: v4 and MVP1 both strong. v4 wins on formatting and the COTS threshold. "
            "Research Optimizer is adequate but thinnest."
        ),
    },
    3: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 4,
        "judgment": (
            "WINNER: Research Optimizer - slight edge over EAGLE v4\n\n"
            "Research Optimizer (Col 5): Exceptional. Both holdings with case facts ($25M min, FY2003/2004 "
            "timing), exact quote ('shall order' language), remedy details, IDIQ-vs-multiyear comparison table. "
            "The deepest legal analysis. Sources: legal-counselor KB files.\n\n"
            "EAGLE v4 (Col 8): Very strong. Both holdings, case facts (ACE contract, IBM, $25M/$5B), direct "
            "quote from GAO, principle table with citations, practical NIH takeaways. Close to Research "
            "Optimizer quality.\n\n"
            "GB MVP1 (Col 7): Good summary of key principles but thinner on case facts. Missing ACE/IBM "
            "specifics, $ amounts. Mentions related precedent B-321640 (useful but not asked). Sources cited "
            "but not fetched.\n\n"
            "EDGE: Research Optimizer and v4 are both excellent. Research Optimizer wins on the 'shall order' "
            "quote and remedy specifics. v4 wins on formatting (table) and NIH-specific practical takeaways."
        ),
    },
    4: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 5,
        "judgment": (
            "WINNER: 3-way tie - all excellent\n\n"
            "Research Optimizer (Col 5): Thorough. GAO Red Book quote, five analytical factors, comparison "
            "table. Strong legal grounding.\n\n"
            "GB MVP1 (Col 7): Very good. Comparison table, governing case B-317139 (FinCEN), specific $ "
            "figures ($2M of $8.98M), Antideficiency Act consequence. Practical and well-structured.\n\n"
            "EAGLE v4 (Col 8): Clear structure. Root statute (31 USC 1502(a)), comparison table, worked "
            "examples with $ figures for both severable and non-severable. Incremental funding explanation.\n\n"
            "EDGE: All three nail this. MVP1 has the strongest case law citation. v4 has the clearest worked "
            "examples. Research Optimizer has the deepest analytical framework. No material gaps in any response."
        ),
    },
    5: {
        "accuracy": 4,
        "completeness": 5,
        "sources": 4,
        "actionability": 4,
        "judgment": (
            "WINNER: EAGLE v4\n\n"
            "CRITICAL: Research Optimizer (Col 5) ANSWERED THE WRONG QUESTION - returned threshold data "
            "instead of fair opportunity exceptions. Major failure (row-shift bug).\n\n"
            "GB MVP1 (Col 7): Correct. Lists 4 exceptions (urgency, unique, follow-on, minimum guarantee). "
            "Notes $7M J&A threshold and approval tiers. Good distinction from Part 6.\n\n"
            "EAGLE v4 (Col 8): Lists 6 exceptions (adds statutory and DoD/NASA/CG Part 6 exception). More "
            "complete than MVP1. Includes J&A approval threshold table. CAVEAT: Cites 'FAR 16.507-6(b)' - "
            "this is the RFO rewrite numbering, not the traditional FAR 16.505(b)(2). Technically correct "
            "under FAC 2025-06 but users familiar with legacy numbering may not recognize it. Deducted 1 "
            "point on accuracy for potential confusion.\n\n"
            "EDGE: v4 is most complete (6 vs 4 exceptions). Research Optimizer is disqualified (wrong "
            "answer). MVP1 is solid but missing 2 exceptions."
        ),
    },
    6: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 5,
        "judgment": (
            "WINNER: GB MVP1 - slight edge\n\n"
            "CRITICAL: Research Optimizer (Col 5) appears to have answered Q5 (fair opportunity) here "
            "instead of Q6 (SBIR). Responses are shifted by one row. Disqualified for this question.\n\n"
            "GB MVP1 (Col 7): Excellent. Starts with jurisdiction threshold (grant vs contract). Three "
            "simultaneous clocks table (debriefing 3-day, GAO 10-day, CICA stay). Critical asymmetry "
            "between FAR 15.505 (pre-award, no tolling) and 15.506 (post-award, tolling). Enhanced "
            "debriefing (DoD DFARS 215.506-70). Very thorough procedural analysis.\n\n"
            "EAGLE v4 (Col 8): Also excellent. Starts with jurisdiction threshold (table of SBIR "
            "structures). Detailed procedural sequence from Day 0. Covers debriefing timing, protest "
            "timeliness, CICA stay interaction. Very thorough.\n\n"
            "EDGE: Both MVP1 and v4 are exceptional on this complex scenario. MVP1 wins on the "
            "three-clocks table clarity and the Day 8 status column. v4 wins on the SBIR structure "
            "threshold table. Both correctly identify the FAR 15.505 no-tolling trap."
        ),
    },
    7: {
        "accuracy": 5,
        "completeness": 4,
        "sources": 3,
        "actionability": 5,
        "judgment": (
            "WINNER: Research Optimizer - slight edge over EAGLE v4\n\n"
            "Research Optimizer (Col 5): Strong design analysis. 'Three walls of the box' metaphor "
            "(budget, timeline, must-have vs nice-to-have). Proposes minimum viable checkpoint set. "
            "Identifies the exit condition problem. Asks follow-up. Actionable.\n\n"
            "GB MVP1 (Col 7): No response.\n\n"
            "EAGLE v4 (Col 8): Good analysis. Identifies dependency chain (Scope -> Labor Mix -> Price -> "
            "Feasibility -> Scope). Three design directions (price-gate, scope card, explicit checkpoint). "
            "Asks good follow-up question. Slightly thinner on the 'why' but more concrete on solutions.\n\n"
            "EDGE: Research Optimizer frames the problem better ('exit condition' insight is key). v4 offers "
            "more concrete solution options. Both are useful. Research Optimizer wins on diagnostic depth; "
            "v4 wins on actionable proposals. No sources cited by either (design discussion, not FAR). "
            "MVP1 did not answer."
        ),
    },
}

for row, r in ratings.items():
    ws.cell(row=row, column=9, value=r["accuracy"]).alignment = wrap
    ws.cell(row=row, column=10, value=r["completeness"]).alignment = wrap
    ws.cell(row=row, column=11, value=r["sources"]).alignment = wrap
    ws.cell(row=row, column=12, value=r["actionability"]).alignment = wrap
    total = r["accuracy"] + r["completeness"] + r["sources"] + r["actionability"]
    ws.cell(row=row, column=13, value=total).alignment = wrap
    ws.cell(row=row, column=14, value=r["judgment"]).alignment = wrap

# Column widths
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 14
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 14
ws.column_dimensions["M"].width = 12
ws.column_dimensions["N"].width = 80

wb.save(XLSX)
print("Done! Wrote ratings to columns I-N")

# Summary
print("\n=== SCORE SUMMARY ===")
print(f"{'Q#':<5} {'Acc':>4} {'Comp':>5} {'Src':>4} {'Act':>4} {'Total':>6}  Winner")
print("-" * 70)
for row in sorted(ratings.keys()):
    r = ratings[row]
    t = r["accuracy"] + r["completeness"] + r["sources"] + r["actionability"]
    winner = r["judgment"].split("\n")[0]
    print(f"Q{row-1:<4} {r['accuracy']:>4} {r['completeness']:>5} {r['sources']:>4} {r['actionability']:>4} {t:>5}/20  {winner}")

totals = [
    r["accuracy"] + r["completeness"] + r["sources"] + r["actionability"]
    for r in ratings.values()
]
avg = sum(totals) / len(totals)
print(f"\n{'AVG':>5} {'':>4} {'':>5} {'':>4} {'':>4} {avg:>5.1f}/20")

# Win/loss tally
print("\n=== WIN/LOSS TALLY ===")
tally = {"EAGLE v4": 0, "Research Optimizer": 0, "GB MVP1": 0, "Tie": 0}
for r in ratings.values():
    w = r["judgment"].split("\n")[0]
    if "EAGLE v4" in w and "tie" not in w.lower():
        tally["EAGLE v4"] += 1
    elif "Research Optimizer" in w:
        tally["Research Optimizer"] += 1
    elif "GB MVP1" in w:
        tally["GB MVP1"] += 1
    elif "tie" in w.lower():
        tally["Tie"] += 1
for k, v in tally.items():
    print(f"  {k}: {v}/{len(ratings)}")
