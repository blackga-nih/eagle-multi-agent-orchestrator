"""Run Jira-ticket-targeted baseline questions against EAGLE and save to Excel.

Each question maps to a specific Jira ticket (EAGLE-70 through EAGLE-77, skipping EAGLE-75 epic).
"""
import asyncio
import json
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

# Row -> (ticket, category, question)
JIRA_QUESTIONS = {
    8: (
        "EAGLE-70",
        "Cascade / KB Search",
        "I need to sole-source a $280,000 annual software maintenance contract to Illumina Inc. "
        "for our BaseSpace Sequence Hub platform. Only Illumina can maintain this proprietary "
        "genomic analysis software. What documents do I need and what's the justification authority?",
    ),
    9: (
        "EAGLE-71",
        "NIH Policy Layer",
        "What are the simplified acquisition threshold and micro-purchase threshold under current FAR? "
        "Also pull any NIH-specific policies that supplement these thresholds, including local purchase "
        "card policies, HCA approval requirements, and NIH procedural steps beyond standard FAR Part 13.",
    ),
    10: (
        "EAGLE-72",
        "GAO KB Retrieval",
        "What did GAO hold in B-302358 regarding IDIQ minimum obligation requirements? "
        "Search the knowledge base for the full decision text and cite the specific KB documents you find.",
    ),
    11: (
        "EAGLE-73",
        "Bona Fide Needs",
        "Explain how the bona fide needs rule under 31 U.S.C. 1502(a) interacts with the "
        "severable vs. non-severable distinction when funding a service contract that crosses "
        "fiscal years. Include both the severable services rules and the time-based bona fide needs rules.",
    ),
    12: (
        "EAGLE-74",
        "FAR 16.505 Exceptions",
        "List every exception to the fair opportunity requirement for task orders under "
        "FAR 16.505(b)(2)(i), including the small business set-aside exception at subparagraph (F). "
        "There should be seven exceptions (A through G). Cite the exact FAR paragraph for each.",
    ),
    13: (
        "EAGLE-76",
        "SBIR KB vs Web",
        "An offeror eliminated from a SBIR Phase II competition wants to protest at GAO. "
        "Search the knowledge base for all relevant protest guidance documents, including "
        "GAO bid protest guides, NIH stay provision policies, and debriefing requirements. "
        "List each KB document found with its file path before analyzing the protest process.",
    ),
    14: (
        "EAGLE-77",
        "Template Path",
        "Generate a Statement of Work for a $200K annual IT help desk support contract for "
        "the NCI Center for Biomedical Informatics and Information Technology (CBIIT). "
        "Requirements: 5 FTEs providing Tier 1-3 support, 24/7 coverage, SLA of 15-minute "
        "response for Tier 1. Before generating, identify which template from the knowledge "
        "base you will use and show its full S3 path.",
    ),
}

SERVER = "http://localhost:8000"
TENANT = "dev-tenant"


async def run_question(client, row, ticket, question):
    session_id = str(uuid.uuid4())
    q_num = row - 1
    print(f"\n{'='*80}")
    print(f"Q{q_num} [{ticket}] (row {row}): {question[:80]}...")
    print(f"Session: {session_id}")
    print(f"{'='*80}")

    start = time.time()
    try:
        resp = await client.post(
            f"{SERVER}/api/chat",
            json={"message": question, "session_id": session_id},
            headers={
                "X-User-Id": "baseline-eval",
                "X-Tenant-Id": TENANT,
                "X-User-Email": "baseline@eval.test",
                "X-User-Tier": "advanced",
            },
            timeout=300.0,
        )
        elapsed = time.time() - start
        data = resp.json()
        response_text = data.get("response", "")
        tools = data.get("tools_called", [])
        usage = data.get("usage", {})
        model = data.get("model", "unknown")

        print(f"\nCompleted in {elapsed:.1f}s | Model: {model}")
        print(f"Tools: {tools}")
        print(f"Response length: {len(response_text):,} chars")
        print(f"\nFirst 500 chars:\n{response_text[:500]}")

        return {
            "row": row, "q_num": q_num, "ticket": ticket,
            "response": response_text, "tools": tools,
            "usage": usage, "model": model,
            "elapsed_s": round(elapsed, 1),
            "session_id": session_id, "status": "ok",
        }
    except Exception as e:
        elapsed = time.time() - start
        print(f"\nERROR after {elapsed:.1f}s: {e}")
        return {
            "row": row, "q_num": q_num, "ticket": ticket,
            "response": f"ERROR: {e}", "tools": [],
            "usage": {}, "model": "error",
            "elapsed_s": round(elapsed, 1),
            "session_id": session_id, "status": "error",
        }


async def main():
    today = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = Path(__file__).resolve().parent.parent / "Use Case List.xlsx"

    print(f"EAGLE V7 Jira Baseline — EAGLE-70 to EAGLE-77")
    print(f"Server: {SERVER}")
    print(f"Excel:  {xlsx_path}")
    print(f"Date:   {today}")
    print(f"Questions: {len(JIRA_QUESTIONS)}")

    # Preflight
    async with httpx.AsyncClient() as client:
        try:
            r = await client.get(f"{SERVER}/api/health", timeout=5)
            health = r.json()
            print(f"Server: {health.get('service', '?')} {health.get('version', '?')} - OK")
        except Exception as e:
            print(f"\nERROR: Server not reachable: {e}")
            sys.exit(1)

    # Add questions to Excel first
    print(f"\nAdding {len(JIRA_QUESTIONS)} Jira questions to Excel rows 8-14...")
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb["Baseline questions"]
    wrap = Alignment(wrap_text=True, vertical="top")

    for row, (ticket, category, question) in JIRA_QUESTIONS.items():
        ws.cell(row=row, column=1, value=row - 1).alignment = wrap  # Q#
        ws.cell(row=row, column=2, value=ticket).alignment = wrap    # Ticket
        ws.cell(row=row, column=3, value=category).alignment = wrap  # Category
        ws.cell(row=row, column=4, value=question).alignment = wrap  # Question
    wb.save(str(xlsx_path))
    print("Questions added.")

    # Run questions
    results = {}
    async with httpx.AsyncClient() as client:
        for row in sorted(JIRA_QUESTIONS.keys()):
            ticket, category, question = JIRA_QUESTIONS[row]
            result = await run_question(client, row, ticket, question)
            results[row] = result

    # Save raw JSON
    json_path = Path(__file__).resolve().parent / "baseline_v7_jira_results.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({str(r): v for r, v in results.items()}, f, indent=2, ensure_ascii=False)
    print(f"\nRaw results saved to {json_path}")

    # Write responses to Excel (same v7 column = 30)
    print(f"\nWriting responses to Excel column AD (col 30)...")
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb["Baseline questions"]
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for row, r in results.items():
        ws.cell(row=row, column=30, value=r["response"]).alignment = wrap

    wb.save(str(xlsx_path))
    print("Done!")

    # Summary
    print(f"\n{'='*90}")
    print("JIRA BASELINE SUMMARY")
    print(f"{'='*90}")
    print(f"{'Q#':<5} {'Ticket':<12} {'Status':<8} {'Time':>6} {'Chars':>8}  Tools")
    print("-" * 90)
    for row in sorted(results.keys()):
        r = results[row]
        status = "OK" if r["status"] == "ok" else "ERROR"
        print(f"Q{r['q_num']:<4} {r['ticket']:<12} {status:<8} {r['elapsed_s']:>5.1f}s {len(r['response']):>7,}  {r['tools']}")

    total_time = sum(r["elapsed_s"] for r in results.values())
    total_chars = sum(len(r["response"]) for r in results.values())
    errors = sum(1 for r in results.values() if r["status"] == "error")
    print(f"\nTotal: {total_time:.0f}s | {total_chars:,} chars | {errors} errors")

    # Jira pass/fail analysis
    print(f"\n{'='*90}")
    print("JIRA TICKET VALIDATION")
    print(f"{'='*90}")
    for row in sorted(results.keys()):
        r = results[row]
        tools = r["tools"]
        ticket = r["ticket"]

        if ticket == "EAGLE-70":
            has_kb = any(t in tools for t in ("knowledge_search", "research", "search_far"))
            web_only = "web_search" in tools and not has_kb
            verdict = "PASS" if has_kb and not web_only else "FAIL"
            print(f"  {ticket}: {verdict} — KB tools: {has_kb}, web-only: {web_only}")

        elif ticket == "EAGLE-71":
            resp_lower = r["response"].lower()
            has_nih = any(x in resp_lower for x in ("nih-specific", "nih specific", "nih policy", "nih manual", "nih faq", "purchase card supplement"))
            print(f"  {ticket}: {'PASS' if has_nih else 'FAIL'} — NIH policy overlay mentioned: {has_nih}")

        elif ticket == "EAGLE-72":
            has_fetch = "knowledge_fetch" in tools or "research" in tools
            resp_has_case = "b-302358" in r["response"].lower() or "b302358" in r["response"].lower()
            print(f"  {ticket}: {'PASS' if has_fetch and resp_has_case else 'PARTIAL'} — KB fetch: {has_fetch}, case cited: {resp_has_case}")

        elif ticket == "EAGLE-73":
            resp_lower = r["response"].lower()
            has_bona_fide = "bona fide" in resp_lower
            has_severable = "severable" in resp_lower
            has_1502 = "1502" in resp_lower
            print(f"  {ticket}: {'PASS' if has_bona_fide and has_severable else 'PARTIAL'} — bona fide: {has_bona_fide}, severable: {has_severable}, 1502: {has_1502}")

        elif ticket == "EAGLE-74":
            resp = r["response"]
            # Count exceptions
            has_7 = any(x in resp for x in ("(G)", "exception (G)", "subparagraph (G)", "seven exceptions", "7 exceptions"))
            has_sb = any(x in resp.lower() for x in ("small business set-aside", "small business set aside", "(f) small business"))
            has_16505 = "16.505" in resp
            has_wrong = "16.507" in resp
            verdict = "PASS" if has_7 and has_sb and has_16505 and not has_wrong else "FAIL"
            print(f"  {ticket}: {verdict} — 7 exceptions: {has_7}, SB set-aside: {has_sb}, 16.505: {has_16505}, wrong 16.507: {has_wrong}")

        elif ticket == "EAGLE-76":
            has_kb = any(t in tools for t in ("knowledge_search", "research", "search_far", "knowledge_fetch"))
            web_before_kb = False  # simplified check
            resp_lower = r["response"].lower()
            has_protest_docs = any(x in resp_lower for x in ("gao_bid_protests", "nih_6033", "protest", "stay provision"))
            print(f"  {ticket}: {'PASS' if has_kb and has_protest_docs else 'PARTIAL'} — KB tools: {has_kb}, protest docs: {has_protest_docs}")

        elif ticket == "EAGLE-77":
            resp_lower = r["response"].lower()
            has_template = any(x in resp_lower for x in ("template", "s3", "eagle-knowledge-base", "sow_template", "approved/"))
            has_sow = any(x in resp_lower for x in ("statement of work", "scope of work", "sow"))
            print(f"  {ticket}: {'PASS' if has_template and has_sow else 'FAIL'} — template path shown: {has_template}, SOW generated: {has_sow}")


if __name__ == "__main__":
    asyncio.run(main())
