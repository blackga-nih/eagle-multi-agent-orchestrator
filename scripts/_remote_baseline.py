"""
_remote_baseline.py — laptop-side driver for running the baseline eval suite
against the DEPLOYED dev/qa backend ALB from inside the VPC, mirroring the
smoke-harness pattern (`_remote_post_deploy_smoke.py`).

Why this exists: the backend ALB is internal-scheme — only resolvable from
inside the VPC. NCI SCP denies `ssm:StartSession`, so port-forward is off the
table. But `ssm:SendCommand` is allowed, so we stage a tiny async runner on
the devbox via base64, run it inside the VPC, and cat the JSON results back.

Round-trip:
  1. Confirm devbox EC2 is running.
  2. Read deployed ALB URL from EagleComputeStack[QA] outputs.
  3. Read selected questions locally from "Use Case List.xlsx" col D.
  4. base64-stage a 30-line async runner on the devbox.
  5. SSM send-command: install httpx, run runner with questions JSON + base URL.
  6. Cat JSON results back over SSM stdout.
  7. Write the responses into the local Excel under a new EAGLE vN column.

Output mirrors `run_baseline.py`:
  - JSON dump at scripts/baseline_{version}_results.json
  - Excel column written with green header
  - Per-question summary table

Usage:
  python scripts/_remote_baseline.py --version v18 --questions=3,5,14,15,16
  python scripts/_remote_baseline.py --version v18 --env qa --questions=3,5,14
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import sys
import time
import uuid
from pathlib import Path

import boto3
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REGION = os.environ.get("AWS_DEFAULT_REGION", "us-east-1")
DEVBOX_STACK = "eagle-ec2-dev"
COMPUTE_STACK_DEV = "EagleComputeStack"
COMPUTE_STACK_QA = "EagleComputeStackQA"
REMOTE_RUNNER = "/tmp/_baseline_runner.py"
REMOTE_QS = "/tmp/_baseline_qs.json"
REMOTE_OUT = "/tmp/_baseline_out.json"
REMOTE_TOKEN = "/tmp/_baseline_token.txt"


def _cognito_login(email: str, password: str, client_id: str) -> str:
    """Mint a Cognito IdToken via USER_PASSWORD_AUTH (mirrors post_deploy_smoke.cognito_login)."""
    cog = boto3.client("cognito-idp", region_name=REGION)
    resp = cog.initiate_auth(
        ClientId=client_id,
        AuthFlow="USER_PASSWORD_AUTH",
        AuthParameters={"USERNAME": email, "PASSWORD": password},
    )
    return resp["AuthenticationResult"]["IdToken"]


# ─────────────────────────────────────────────────────────────────
# Embedded runner — runs ON the devbox via SSM
# ─────────────────────────────────────────────────────────────────
RUNNER_SOURCE = '''
import asyncio
import json
import os
import sys
import uuid
import time

import httpx


async def run_one(client: httpx.AsyncClient, base: str, q: dict, tenant: str, token: str) -> dict:
    """POST /api/chat/stream and accumulate text events.

    REST /api/chat buffers the full agent response (5-8 min) and trips the
    ALB response timeout (~200s) → HTTP 502. SSE keeps the connection alive
    by streaming events, dodging the timeout. Same pattern as smoke harness.

    Each call gets a fresh session_id AND a fresh user_id so the deployed
    backend cannot share any per-tenant/user state across questions
    (workspace cache, package context, in-memory KB depth tracker, etc.).
    Session storage is keyed by (tenant_id, user_id, session_id), so a new
    user_id per call guarantees a clean slate even if anything upstream
    inadvertently keys by user.
    """
    sid = str(uuid.uuid4())
    uid = f"baseline-q{q['q_num']}-{uuid.uuid4().hex[:8]}"
    start = time.time()
    headers = {
        "X-User-Id": uid,
        "X-Tenant-Id": tenant,
        "X-User-Email": f"{uid}@eval.test",
        "X-User-Tier": "advanced",
        "Accept": "text/event-stream",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = {"message": q["question"], "session_id": sid}
    url = f"{base}/api/chat/stream"

    text_chunks: list[str] = []
    tools: list[str] = []
    usage: dict = {}
    model = "unknown"
    final_text = ""
    status = 0
    error_msg = ""
    # research tool_result payloads — each carries lane_breakdown + per-doc info.
    # Multiple research events can fire (supervisor + subagent); keep them all
    # and let the post-processor pick the richest one.
    research_results: list[dict] = []
    source_docs: list[str] = []  # all unique doc paths surfaced by any tool_result

    try:
        async with client.stream("POST", url, json=body, headers=headers, timeout=600.0) as resp:
            status = resp.status_code
            if not resp.is_success:
                # Read body for error detail
                async for chunk in resp.aiter_bytes():
                    error_msg += chunk.decode("utf-8", errors="ignore")
                    if len(error_msg) > 500:
                        break
                elapsed = time.time() - start
                return {
                    "row": q["row"], "q_num": q["q_num"], "status": "error",
                    "response": f"HTTP {status}: {error_msg.strip()[:300]}",
                    "tools": [], "usage": {}, "model": "error",
                    "elapsed_s": round(elapsed, 1), "session_id": sid, "user_id": uid,
                }
            async for line in resp.aiter_lines():
                if not line or not line.startswith("data:"):
                    continue
                payload = line[5:].strip()
                if not payload or payload == "[DONE]":
                    continue
                try:
                    e = json.loads(payload)
                except json.JSONDecodeError:
                    continue
                et = e.get("type") or e.get("event") or ""
                if et == "text":
                    t = e.get("text") or e.get("content") or ""
                    if isinstance(t, str):
                        text_chunks.append(t)
                elif et == "tool_use":
                    name = e.get("name") or (e.get("tool_use") or {}).get("name")
                    if name and name not in tools:
                        tools.append(name)
                elif et == "tool_result":
                    # Per PR #161 (cf0d3bb): research tool_result carries the
                    # lane_breakdown + per-doc summary. The agent no longer
                    # inlines KB paths in the response text — sources only
                    # surface here.
                    tr = e.get("tool_result") or {}
                    name = tr.get("name") or e.get("name") or ""
                    res = tr.get("result")
                    if isinstance(res, str):
                        try:
                            res = json.loads(res)
                        except json.JSONDecodeError:
                            res = {"_raw": res}
                    if name == "research" and isinstance(res, dict):
                        research_results.append(res)
                    # Walk any structure for paths that look like KB doc keys
                    def _walk(obj):
                        if isinstance(obj, dict):
                            for v in obj.values():
                                _walk(v)
                        elif isinstance(obj, list):
                            for v in obj:
                                _walk(v)
                        elif isinstance(obj, str):
                            for tok in obj.split():
                                tok = tok.strip("`,;:()[]<>'\\\"")
                                if "/" in tok and any(tok.endswith(ext) for ext in (".txt", ".docx", ".pdf", ".md", ".xlsx", ".doc", ".html", ".json")):
                                    if tok.startswith("eagle-knowledge-base/"):
                                        tok = tok.split("eagle-knowledge-base/approved/", 1)[-1]
                                    if tok and tok not in source_docs:
                                        source_docs.append(tok)
                    _walk(res)
                elif et == "complete":
                    # complete may carry the consolidated text + usage
                    ft = e.get("text") or e.get("response")
                    if isinstance(ft, str) and ft:
                        final_text = ft
                    u = e.get("usage")
                    if isinstance(u, dict):
                        usage = u
                    m = e.get("model")
                    if isinstance(m, str):
                        model = m
                elif et == "error":
                    msg = e.get("error") or e.get("message") or json.dumps(e)[:200]
                    elapsed = time.time() - start
                    return {
                        "row": q["row"], "q_num": q["q_num"], "status": "error",
                        "response": f"SSE error event: {msg}",
                        "tools": tools, "usage": usage, "model": model,
                        "elapsed_s": round(elapsed, 1), "session_id": sid, "user_id": uid,
                    }
        elapsed = time.time() - start
        response_text = final_text or "".join(text_chunks)
        if source_docs and response_text:
            sep = chr(10)
            footer_lines = ["", "## Sources"]
            for d in source_docs:
                footer_lines.append(f"- `eagle-knowledge-base/approved/{d}`")
            response_text = response_text + sep + sep.join(footer_lines)
        return {
            "row": q["row"], "q_num": q["q_num"],
            "status": "ok" if response_text else "error",
            "response": response_text or "(empty SSE stream)",
            "tools": tools, "usage": usage, "model": model,
            "elapsed_s": round(elapsed, 1), "session_id": sid, "user_id": uid,
            "source_docs": source_docs,
            "research_results": research_results,
        }
    except Exception as e:
        elapsed = time.time() - start
        return {
            "row": q["row"], "q_num": q["q_num"], "status": "error",
            "response": f"ERROR: {type(e).__name__}: {e}",
            "tools": tools, "usage": usage, "model": "error",
            "elapsed_s": round(elapsed, 1), "session_id": sid, "user_id": uid,
        }


async def main() -> int:
    qs_path, base_url, tenant, out_path = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]
    token_path = sys.argv[5] if len(sys.argv) > 5 else ""
    token = ""
    if token_path and os.path.exists(token_path):
        with open(token_path) as f:
            token = f.read().strip()
    with open(qs_path) as f:
        questions = json.load(f)
    auth_note = "with Bearer token" if token else "no token"
    print(f"Running {len(questions)} questions SEQUENTIALLY against {base_url} "
          f"(tenant={tenant}, {auth_note}). Each gets a fresh session_id + user_id.",
          flush=True)
    # Sequential execution + fresh AsyncClient per question — no shared
    # connection pool, no concurrent in-flight requests against the deployed
    # backend. Combined with per-question user_id (set in run_one), this
    # eliminates any cross-question state on the server.
    results: list[dict] = []
    for q in questions:
        async with httpx.AsyncClient() as client:
            r = await run_one(client, base_url, q, tenant, token)
        results.append(r)
        print(f"  done Q{r['q_num']:>3} {r['status']:<6} {r['elapsed_s']:>6.1f}s  "
              f"{len(r.get('response', '')):>6} chars  "
              f"sid={r.get('session_id', '')[:8]} uid={r.get('user_id', '')}",
              flush=True)
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)
    for r in results:
        print(f"  Q{r['q_num']:>3} {r['status']:<6} {r['elapsed_s']:>6.1f}s  "
              f"{len(r.get('response', '')):>6} chars  tools={r.get('tools', [])}",
              flush=True)
    print(f"Wrote {len(results)} results to {out_path}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
'''


# ─────────────────────────────────────────────────────────────────
# AWS helpers (lifted verbatim from _remote_post_deploy_smoke.py)
# ─────────────────────────────────────────────────────────────────
def _cfn_outputs(stack_name: str) -> dict[str, str]:
    cf = boto3.client("cloudformation", region_name=REGION)
    desc = cf.describe_stacks(StackName=stack_name)["Stacks"][0]
    return {o["OutputKey"]: o["OutputValue"] for o in desc.get("Outputs", [])}


def _devbox_instance_id() -> str:
    return _cfn_outputs(DEVBOX_STACK)["InstanceId"]


def _ensure_running(iid: str) -> None:
    ec2 = boto3.client("ec2", region_name=REGION)
    state = ec2.describe_instances(InstanceIds=[iid])["Reservations"][0]["Instances"][0]["State"]["Name"]
    if state == "running":
        return
    if state != "stopped":
        sys.exit(f"Devbox is in state '{state}' — wait or fix manually.")
    print(f"Starting devbox {iid}...")
    ec2.start_instances(InstanceIds=[iid])
    ec2.get_waiter("instance_running").wait(InstanceIds=[iid])
    print("Devbox started; waiting 15s for SSM agent...")
    time.sleep(15)


def _ssm_run(iid: str, commands: list[str], comment: str, timeout: int = 600) -> dict:
    ssm = boto3.client("ssm", region_name=REGION)
    cmd = ssm.send_command(
        InstanceIds=[iid],
        DocumentName="AWS-RunShellScript",
        Parameters={"commands": commands, "executionTimeout": [str(timeout)]},
        Comment=comment[:100],
        TimeoutSeconds=timeout + 60,
    )
    cid = cmd["Command"]["CommandId"]
    while True:
        time.sleep(3)
        r = ssm.get_command_invocation(CommandId=cid, InstanceId=iid)
        if r["Status"] in ("Success", "Failed", "Cancelled", "TimedOut"):
            return r


def _print_block(title: str, body: str, max_chars: int = 4000) -> None:
    print(f"--- {title} ---")
    if not body:
        print("(empty)")
        return
    if len(body) > max_chars:
        print(body[:max_chars])
        print(f"... [{len(body) - max_chars} chars truncated]")
    else:
        print(body)


# ─────────────────────────────────────────────────────────────────
# Question selection (from local Excel)
# ─────────────────────────────────────────────────────────────────
def _parse_question_spec(spec: str) -> list[int]:
    nums: list[int] = []
    for part in spec.split(","):
        part = part.strip().lstrip("qQ")
        if "-" in part:
            a, b = part.split("-", 1)
            nums.extend(range(int(a), int(b) + 1))
        elif part:
            nums.append(int(part))
    return nums


def _read_questions(xlsx_path: Path, q_nums: list[int] | None) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Baseline questions"]
    questions: list[dict] = []
    for row in range(2, ws.max_row + 1):
        text = ws.cell(row=row, column=4).value
        if not text:
            continue
        q_num = row - 1
        if q_nums and q_num not in q_nums:
            continue
        questions.append({"row": row, "q_num": q_num, "question": str(text)})
    return questions


# ─────────────────────────────────────────────────────────────────
# Excel write-back (mirrors run_baseline.py output formatting)
# ─────────────────────────────────────────────────────────────────
def _write_to_excel(xlsx_path: Path, results: list[dict], version: str) -> str:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Baseline questions"]
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        col += 1
    today = time.strftime("%Y-%m-%d")
    header_cell = ws.cell(row=1, column=col, value=f"EAGLE {version.upper()} Response ({today})")
    header_cell.fill = PatternFill("solid", fgColor="2E7D32")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in results:
        cell = ws.cell(row=r["row"], column=col, value=r.get("response", ""))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 80
    wb.save(xlsx_path)
    return openpyxl.utils.get_column_letter(col)


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────
def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--version", required=True, help="Version label (e.g. v18)")
    p.add_argument("--env", default="dev", choices=["dev", "qa"])
    p.add_argument("--questions", default="", help="e.g. 3,5,14,15,16 or 7-10. Empty = all.")
    p.add_argument("--tenant", default="dev-tenant")
    p.add_argument("--xlsx", default="Use Case List.xlsx")
    p.add_argument("--timeout", type=int, default=3600,
                   help="Per-step SSM timeout (seconds). Default 3600s = 60 min "
                        "(sequential mode runs questions one-by-one to avoid "
                        "any cross-question backend state, so total wall time "
                        "scales with question count × ~5min each).")
    p.add_argument("--auth", action="store_true",
                   help="Mint a Cognito IdToken (requires EAGLE_TEST_EMAIL, EAGLE_TEST_PASSWORD, "
                        "COGNITO_CLIENT_ID env vars). Required for deployed dev/qa.")
    args = p.parse_args()

    token = ""
    if args.auth:
        email = os.environ.get("EAGLE_TEST_EMAIL")
        password = os.environ.get("EAGLE_TEST_PASSWORD")
        client_id = os.environ.get("COGNITO_CLIENT_ID")
        if not (email and password and client_id):
            sys.exit("--auth requires EAGLE_TEST_EMAIL, EAGLE_TEST_PASSWORD, COGNITO_CLIENT_ID env vars")
        print(f"Minting Cognito IdToken for {email}...")
        token = _cognito_login(email, password, client_id)
        print(f"    token acquired ({len(token)} chars)")

    repo_root = Path(__file__).resolve().parent.parent
    xlsx_path = (repo_root / args.xlsx).resolve()
    if not xlsx_path.exists():
        sys.exit(f"Missing xlsx: {xlsx_path}")

    q_nums = _parse_question_spec(args.questions) if args.questions else None
    questions = _read_questions(xlsx_path, q_nums)
    if not questions:
        sys.exit(f"No questions matched spec {args.questions!r}")
    print(f"Selected {len(questions)} questions: " + ", ".join(f"Q{q['q_num']}" for q in questions))

    iid = _devbox_instance_id()
    _ensure_running(iid)

    compute_stack = COMPUTE_STACK_QA if args.env == "qa" else COMPUTE_STACK_DEV
    outs = _cfn_outputs(compute_stack)
    backend_url = outs["BackendUrl"].rstrip("/")
    print(f"=== {args.env} backend = {backend_url}")

    # Stage runner + questions JSON + (optional) token via base64
    runner_b64 = base64.b64encode(RUNNER_SOURCE.encode()).decode()
    qs_json = json.dumps(questions)
    qs_b64 = base64.b64encode(qs_json.encode()).decode()
    token_b64 = base64.b64encode(token.encode()).decode() if token else ""

    stage_cmds = [
        f"echo {runner_b64} | base64 -d > {REMOTE_RUNNER}",
        f"echo {qs_b64} | base64 -d > {REMOTE_QS}",
        "python3 -m pip install --quiet 'httpx>=0.27' 2>&1 | tail -3",
        f"wc -l {REMOTE_RUNNER} && wc -c {REMOTE_QS}",
    ]
    if token_b64:
        stage_cmds.insert(2, f"echo {token_b64} | base64 -d > {REMOTE_TOKEN}")

    print(f"\n=== Stage runner ({len(runner_b64)} b64 chars) + questions ({len(qs_b64)} b64 chars)"
          + (f" + token ({len(token_b64)} b64 chars)" if token_b64 else "") + " ===")
    stage = _ssm_run(
        iid,
        stage_cmds,
        "remote-baseline: stage",
        timeout=180,
    )
    _print_block("stage stdout", stage.get("StandardOutputContent", ""))
    if stage["Status"] != "Success":
        _print_block("stage stderr", stage.get("StandardErrorContent", ""))
        return 1

    print(f"\n=== Run baseline against {args.env} backend ALB (timeout {args.timeout}s) ===")
    token_arg = REMOTE_TOKEN if token else '""'
    run = _ssm_run(
        iid,
        [
            f"python3 {REMOTE_RUNNER} {REMOTE_QS} {backend_url} {args.tenant} {REMOTE_OUT} {token_arg}",
        ],
        f"remote-baseline: run {args.version} on {args.env}",
        timeout=args.timeout,
    )
    _print_block("run stdout", run.get("StandardOutputContent", ""), max_chars=8000)
    if run["Status"] != "Success":
        _print_block("run stderr", run.get("StandardErrorContent", ""))
        return 1

    print(f"\n=== Cat results back from devbox (gzip+b64 to dodge SSM stdout truncation) ===")
    cat = _ssm_run(
        iid,
        [f"gzip -c {REMOTE_OUT} | base64 -w0"],
        "remote-baseline: cat results",
        timeout=60,
    )
    if cat["Status"] != "Success":
        _print_block("cat stderr", cat.get("StandardErrorContent", ""))
        return 1
    raw_b64 = cat.get("StandardOutputContent", "").strip()
    import gzip
    try:
        compressed = base64.b64decode(raw_b64)
        raw = gzip.decompress(compressed).decode("utf-8")
        results = json.loads(raw)
    except (ValueError, gzip.BadGzipFile, json.JSONDecodeError) as exc:
        print(f"Failed to parse results JSON: {exc}")
        print(f"raw_b64 len: {len(raw_b64)}; first 500: {raw_b64[:500]}")
        return 1

    # Local outputs
    out_json = repo_root / "scripts" / f"baseline_{args.version}_dev_results.json"
    out_json.write_text(json.dumps(results, indent=2), encoding="utf-8")
    print(f"\nResults JSON saved to {out_json}")

    col_letter = _write_to_excel(xlsx_path, results, args.version)
    print(f"Excel updated: {xlsx_path} (column {col_letter})")

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"{'Q#':<6} {'Status':<8} {'Time':<10} {'Chars':<8} Tools")
    print("-" * 80)
    for r in results:
        tools = ",".join(r.get("tools", [])) or "-"
        print(f"Q{r['q_num']:<5} {r['status']:<8} {r.get('elapsed_s', 0):<10.1f} "
              f"{len(r.get('response', '')):<8} {tools}")
    errors = sum(1 for r in results if r["status"] != "ok")
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
