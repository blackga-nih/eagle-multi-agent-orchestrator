"""Judge EAGLE v5 baseline responses against v4 and prior baselines."""
import openpyxl
import sys
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

XLSX = "C:/Users/blackga/Desktop/eagle/sm_eagle/Use Case List.xlsx"
wb = openpyxl.load_workbook(XLSX)
ws = wb["Baseline questions"]

# New columns for v5 scoring — columns 17-22 (Q-V)
COLS = {
    17: "EAGLE v5 Accuracy (0-5)",
    18: "EAGLE v5 Completeness (0-5)",
    19: "EAGLE v5 Sources (0-5)",
    20: "EAGLE v5 Actionability (0-5)",
    21: "EAGLE v5 Total (0-20)",
    22: "v5 vs v4 Comparative Judgment",
}

wrap = Alignment(wrap_text=True, vertical="top")
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, label in COLS.items():
    cell = ws.cell(row=1, column=col_num, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

ratings = {
    2: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 4,
        "actionability": 5,
        "judgment": (
            "v5 = v4 (no change)\n\n"
            "Both versions produce identical quality. Same table format, same threshold "
            "values, same nuances (construction $2K, SCA $2.5K, contingency $25K). v5 adds "
            "explicit practical implications section (below $15K / $15K-$350K / above $350K) "
            "which is marginally more actionable.\n\n"
            "Tools used: v4=query_compliance_matrix | v5=query_compliance_matrix\n\n"
            "This is a threshold lookup — the compliance matrix tool handles it well in both "
            "versions. No KB cascade effect expected or observed.\n\n"
            "VERDICT: Tie. Both excellent."
        ),
    },
    3: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 5,
        "judgment": (
            "v5 > v4 (improved)\n\n"
            "KEY IMPROVEMENT: v5 now includes the exact 'partial homonymy' quote from the "
            "GAO decision — v4 had a DIFFERENT quote ('Recording evidences the obligation but "
            "does not create it') suggesting model knowledge rather than primary source. v5's "
            "quote matches the actual decision text.\n\n"
            "v5 also adds:\n"
            "  - CBP/DHS IG request context (31 USC 3529)\n"
            "  - The erroneous FAR 52.217-2 insertion detail\n"
            "  - Explicit obligation framework table (award vs task order)\n"
            "  - Clean practical takeaways section for NIH COs\n"
            "  - Sources section citing GAO Red Book Vol II Ch 7\n\n"
            "Tools used: v4=knowledge_search, web_search | v5=knowledge_search, web_search, legal_counsel\n\n"
            "v5 called legal_counsel subagent which deepened the analysis. The response now "
            "reads like it was sourced from the actual decision, not reconstructed from model "
            "knowledge.\n\n"
            "SOURCE GAP STATUS: The v4 gap was 'used a non-verbatim quote suggesting model "
            "knowledge.' v5 now includes the verbatim 'partial homonymy' language from the "
            "decision. GAP PARTIALLY CLOSED.\n\n"
            "VERDICT: v5 wins. Deeper source grounding, better structured, more actionable."
        ),
    },
    4: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 5,
        "judgment": (
            "v5 > v4 (improved)\n\n"
            "KEY IMPROVEMENT: v5 now includes the GAO Red Book direct quote ('Is the contract "
            "essentially for a single undertaking, or does it envision a series of separate "
            "tasks?') — this was the GAP identified in v4's source analysis. v4 paraphrased "
            "the concept; v5 quotes it.\n\n"
            "v5 also adds:\n"
            "  - 'Short Bridge Exception' section (1-3 month overlap)\n"
            "  - Common Gray Areas table (agile dev, O&M, pen test, SOC, PM)\n"
            "  - 'Critical Errors to Avoid' section (3 common misconceptions)\n"
            "  - Antideficiency Act exposure warning\n"
            "  - More case citations (B-223096, B-237407, B-235580)\n\n"
            "Tools used: v4=knowledge_search | v5=knowledge_search, knowledge_fetch\n\n"
            "THE CASCADE EFFECT: v5 actually called knowledge_fetch after knowledge_search. "
            "This is the exact behavior the KB cascade enforcement was designed to produce. "
            "v4 answered from search summaries; v5 read the full document.\n\n"
            "SOURCE GAP STATUS: v4 gap was 'paraphrased rather than quoting the Red Book.' "
            "v5 now includes the verbatim quote. GAP CLOSED.\n\n"
            "VERDICT: v5 wins. Deeper, better structured, quotes primary source."
        ),
    },
    5: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 5,
        "judgment": (
            "v5 > v4 (improved)\n\n"
            "KEY IMPROVEMENT: v5 used search_far + knowledge_fetch. The response includes "
            "explicit source citation to FAR_Part_16_IDIQ_Comprehensive_RFO_2025.txt — the "
            "71K-char file that was the primary GAP identified in v4's source analysis.\n\n"
            "v5 now includes:\n"
            "  - J&A approval threshold table (scaled by dollar value) — v4 had this\n"
            "  - Posting requirement details (14-day / 30-day) — NEW in v5\n"
            "  - Small business set-aside exemption note — NEW in v5\n"
            "  - DoD/NASA/Coast Guard threshold differences — NEW in v5\n"
            "  - Explicit source KB file path — NEW in v5\n\n"
            "Tools used: v4=search_far | v5=search_far, knowledge_fetch\n\n"
            "THE CASCADE EFFECT: v5 called knowledge_fetch after search_far. This pulled the "
            "full RFO text and allowed the agent to include posting requirements and the "
            "Part 6 distinction that v4 lacked.\n\n"
            "SOURCE GAP STATUS: v4 gap was 'missing 9-item J&A content checklist and $10M "
            "protest threshold.' v5 doesn't enumerate the 9-item checklist explicitly but "
            "includes the posting and approval framework. The protest threshold for civilian "
            "agencies is not mentioned. GAP PARTIALLY CLOSED.\n\n"
            "VERDICT: v5 wins on source depth. Both have the 6 exceptions correct."
        ),
    },
    6: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 5,
        "actionability": 5,
        "judgment": (
            "v5 >= v4 (comparable, slight edge v5)\n\n"
            "Both v4 and v5 produce exceptional responses to this complex SBIR protest "
            "scenario. Both are 13-15K chars of deep procedural analysis.\n\n"
            "KEY IMPROVEMENT: v5 cites specific KB file paths:\n"
            "  - GAO_Bid_Protests_Essential_Guide_for_COs.txt\n"
            "  - NIH_Policy_6033_2_Protest_Stay_Provisions_2008.txt\n"
            "  - NIH_6033_1_GAO_Protest_CO_Statement.txt\n"
            "  - FC_33_Protests_Disputes_Appeals.txt\n"
            "  - GAO_B-423341_ECP_Misleading_Discussions_Best_Value.txt\n"
            "These are explicit KB source references — v4 cited the same documents "
            "conceptually but not as KB file paths.\n\n"
            "Tools used: v4=knowledge_search, search_far, knowledge_fetch | "
            "v5=knowledge_search, search_far, knowledge_fetch\n\n"
            "Both versions called knowledge_fetch. The cascade enforcement didn't change "
            "behavior here because v4 already fetched documents for this complex question.\n\n"
            "v5 structural improvements:\n"
            "  - Cleaner section numbering (I-VIII)\n"
            "  - Summary Decision Matrix at the end\n"
            "  - Action Items section with 6 specific steps\n"
            "  - Timeline diagram with Day numbers\n"
            "  - NIH Manual 6033-2 override D&F requirements table\n\n"
            "SOURCE GAP STATUS: v4 gaps were (1) J&A Desk Guide debriefing disclosure rules "
            "and (2) NIH 6315-1 SBIR peer review documentation requirements. v5 now cites "
            "GAO Essential Guide and NIH 6033-2 directly. The J&A Desk Guide is still not "
            "fetched explicitly, but the debriefing disclosure rules are covered via FAR "
            "15.505(e). SBIR peer review documentation (NIH 6315-1) is still not addressed. "
            "GAP PARTIALLY CLOSED (1 of 2).\n\n"
            "VERDICT: v5 slight edge on source citations and structure. Both excellent."
        ),
    },
    7: {
        "accuracy": 5,
        "completeness": 5,
        "sources": 3,
        "actionability": 5,
        "judgment": (
            "v5 > v4 (improved)\n\n"
            "This is a design discussion — no regulatory sources expected. Both versions "
            "correctly identify the core tension (speed vs. completeness) and propose "
            "solutions.\n\n"
            "v5 improvements over v4:\n"
            "  - Three Gates framework (dollar range, what's bought, existing vehicle)\n"
            "  - Specific 'backwards problem' analysis with 3 solution options\n"
            "  - 'Rough IGCE before detailed SOW' recommendation — a specific, actionable "
            "    proposal that directly addresses the user's pain point\n"
            "  - Clearer table (entry point, what Eagle did, what broke)\n"
            "  - Ends with a focused follow-up question\n\n"
            "v4 was good but more diagnostic than prescriptive. v5 offers concrete next "
            "steps.\n\n"
            "Tools used: v4=none | v5=none\n"
            "No cascade effect — design discussion, not a regulatory question.\n\n"
            "VERDICT: v5 wins on actionability. Better structured proposals."
        ),
    },
}

for row, r in ratings.items():
    ws.cell(row=row, column=17, value=r["accuracy"]).alignment = wrap
    ws.cell(row=row, column=18, value=r["completeness"]).alignment = wrap
    ws.cell(row=row, column=19, value=r["sources"]).alignment = wrap
    ws.cell(row=row, column=20, value=r["actionability"]).alignment = wrap
    total = r["accuracy"] + r["completeness"] + r["sources"] + r["actionability"]
    ws.cell(row=row, column=21, value=total).alignment = wrap
    ws.cell(row=row, column=22, value=r["judgment"]).alignment = wrap

# Column widths
ws.column_dimensions["Q"].width = 12
ws.column_dimensions["R"].width = 14
ws.column_dimensions["S"].width = 12
ws.column_dimensions["T"].width = 14
ws.column_dimensions["U"].width = 12
ws.column_dimensions["V"].width = 100

wb.save(XLSX)
print("Done! Wrote v5 ratings to columns Q-V\n")

# Summary
print("=" * 90)
print("EAGLE v5 SCORE SUMMARY")
print("=" * 90)
print(f"{'Q#':<5} {'Acc':>4} {'Comp':>5} {'Src':>4} {'Act':>4} {'Total':>6}  v4 Total  Delta  Verdict")
print("-" * 90)

# v4 scores for comparison
v4_scores = {
    2: {"accuracy": 5, "completeness": 5, "sources": 4, "actionability": 4},
    3: {"accuracy": 5, "completeness": 5, "sources": 5, "actionability": 4},
    4: {"accuracy": 5, "completeness": 5, "sources": 5, "actionability": 5},
    5: {"accuracy": 4, "completeness": 5, "sources": 4, "actionability": 4},
    6: {"accuracy": 5, "completeness": 5, "sources": 5, "actionability": 5},
    7: {"accuracy": 5, "completeness": 4, "sources": 3, "actionability": 5},
}

for row in sorted(ratings.keys()):
    r = ratings[row]
    v5_total = r["accuracy"] + r["completeness"] + r["sources"] + r["actionability"]
    v4 = v4_scores[row]
    v4_total = v4["accuracy"] + v4["completeness"] + v4["sources"] + v4["actionability"]
    delta = v5_total - v4_total
    delta_str = f"+{delta}" if delta > 0 else str(delta)
    verdict = r["judgment"].split("\n")[0]
    print(f"Q{row-1:<4} {r['accuracy']:>4} {r['completeness']:>5} {r['sources']:>4} {r['actionability']:>4} {v5_total:>5}/20  {v4_total:>5}/20  {delta_str:>5}  {verdict}")

v5_totals = [
    r["accuracy"] + r["completeness"] + r["sources"] + r["actionability"]
    for r in ratings.values()
]
v4_totals = [
    v["accuracy"] + v["completeness"] + v["sources"] + v["actionability"]
    for v in v4_scores.values()
]
v5_avg = sum(v5_totals) / len(v5_totals)
v4_avg = sum(v4_totals) / len(v4_totals)
delta_avg = v5_avg - v4_avg

print(f"\n{'AVG':>5} {'':>4} {'':>5} {'':>4} {'':>4} {v5_avg:>5.1f}/20  {v4_avg:>5.1f}/20  {delta_avg:>+5.1f}")

# Win/loss tally
print("\n" + "=" * 90)
print("v5 vs v4 TALLY")
print("=" * 90)
v5_wins = 0
ties = 0
v4_wins = 0
for r in ratings.values():
    verdict = r["judgment"].split("\n")[0].lower()
    if "v5 > v4" in verdict:
        v5_wins += 1
    elif "v5 >= v4" in verdict:
        v5_wins += 1  # slight edge counts
    elif "v5 = v4" in verdict or "no change" in verdict or "tie" in verdict:
        ties += 1
    else:
        v4_wins += 1

print(f"  v5 wins:  {v5_wins}/{len(ratings)}")
print(f"  Ties:     {ties}/{len(ratings)}")
print(f"  v4 wins:  {v4_wins}/{len(ratings)}")

# Source gap closure
print("\n" + "=" * 90)
print("KB CASCADE ENFORCEMENT IMPACT")
print("=" * 90)
print("""
Q1 (Thresholds): No cascade effect — compliance matrix tool handles this.
Q2 (GAO B-302358): v5 now includes verbatim 'partial homonymy' quote. GAP PARTIALLY CLOSED.
Q3 (Severable): v5 called knowledge_fetch after search. Includes Red Book quote. GAP CLOSED.
Q4 (Fair Opportunity): v5 called knowledge_fetch after search_far. Source cited. GAP PARTIALLY CLOSED.
Q5 (SBIR Protest): v5 cites 5 KB files by path. J&A Desk Guide gap partially addressed. GAP PARTIALLY CLOSED.
Q6 (Design): No cascade effect — design discussion.

TOOLS COMPARISON:
  Q3: v4=knowledge_search only -> v5=knowledge_search + knowledge_fetch  ** CASCADE WORKING **
  Q4: v4=search_far only       -> v5=search_far + knowledge_fetch        ** CASCADE WORKING **
  Q5: v4=all three              -> v5=all three (already fetching)        (no change needed)

The _fetch_reminder injection is producing the desired effect: on Q3 and Q4, the agent
now fetches full documents after search, where v4 answered from summaries alone.
""")
