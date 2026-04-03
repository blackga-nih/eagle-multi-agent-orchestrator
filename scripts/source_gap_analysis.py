"""Source gap analysis: RO sources missing in EAGLE v4."""
import openpyxl
import sys
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

XLSX = "C:/Users/blackga/Desktop/eagle/sm_eagle/Use Case List.xlsx"
wb = openpyxl.load_workbook(XLSX)
ws = wb["Baseline questions"]

col = 15
cell = ws.cell(row=1, column=col, value="Source Gaps: RO Sources Missing in EAGLE v4")
cell.font = Font(bold=True, color="FFFFFF")
cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
cell.alignment = Alignment(wrap_text=True, vertical="top")

wrap = Alignment(wrap_text=True, vertical="top")

gaps = {}

gaps[2] = (
    "NO ESSENTIAL SOURCE GAPS\n\n"
    "RO sourced: compliance-strategist/PMR-checklists/HHS_PMR_Threshold_Matrix.txt (2,589 chars)\n"
    "EAGLE v4 sourced: query_compliance_matrix tool (matrix.json thresholds)\n\n"
    "Both arrive at identical core facts. v4 actually surfaces MORE detail (construction, SCA, "
    "contingency, COTS exceptions) without needing the KB doc. The compliance matrix tool "
    "encodes the same data RO pulled from the HHS PMR Threshold Matrix file.\n\n"
    "VERDICT: No gap. v4 tool-based approach is equivalent or better."
)

gaps[3] = (
    "TWO SOURCE GAPS IDENTIFIED\n\n"
    "RO fetched TWO KB documents that EAGLE v4 did not:\n\n"
    "GAP 1: legal-counselor/appropriations-law/GAO_B-302358_IDIQ_Min_Fund.txt (28,370 chars)\n"
    "  This is the FULL GAO decision text. RO read it and extracted:\n"
    "  - The exact \"shall order\" quote from FAR 52.216-22\n"
    "  - Specific dollar amounts: $45K from FY2003, $955K from FY2004\n"
    "  - The \"partial homonymy\" language re: multiyear vs multiple-year\n"
    "  - Remedies: deobligate $955K, re-obligate from FY2003, ADA determination\n"
    "  EAGLE v4 DID reproduce both holdings, case facts (ACE, IBM, $25M/$5B), "
    "and a GAO quote -- but a DIFFERENT quote (\"Recording evidences the obligation "
    "but does not create it\") suggesting it used model knowledge, not the primary source.\n"
    "  IMPACT: Medium. v4 is substantively correct but lacks primary source grounding. "
    "A CO checking the quote may not find it verbatim in the decision.\n\n"
    "GAP 2: financial-advisor/appropriations-law/appropriations_law_IDIQ_funding.txt (18,283 chars)\n"
    "  Broader IDIQ funding doctrine file. RO surfaced companion cases:\n"
    "  - B-321640 (\"parking funds\" violation)\n"
    "  - B-308969 (interagency IDIQ funding)\n"
    "  - The interagency obligation rule (both agencies must obligate)\n"
    "  v4 cited B-280945 (nominal minimum) instead -- valid but different.\n"
    "  IMPACT: Low-medium. Companion cases are useful context but not asked for.\n\n"
    "VERDICT: RO has deeper primary source grounding. v4 is factually correct but would "
    "benefit from fetching the full decision text."
)

gaps[4] = (
    "ONE MINOR SOURCE GAP\n\n"
    "RO fetched: appropriations_law_severable_services.txt (17,549 chars, same file at two paths)\n\n"
    "RO extracted that v4 did NOT include:\n"
    "  - The GAO Red Book direct quote (\"The question is whether the contract is essentially "
    "for a single undertaking or job...\")\n"
    "  - Five named analytical factors as a framework\n\n"
    "v4 DID cite the same KB file in Sources and covered equivalent ground with:\n"
    "  - Comparison table + worked examples\n"
    "  - MORE case law citations (B-223096, B-237407, B-235580, B-288266)\n"
    "  - FAR clause references (52.232-19, 52.232-22)\n\n"
    "VERDICT: Minor gap. v4 cited the file but paraphrased rather than quoting. "
    "Compensated with more case citations and clearer examples."
)

gaps[5] = (
    "RO ANSWERED THE WRONG QUESTION -- SHIFTED ROW BUG\n\n"
    "RO returned threshold data here instead of fair opportunity exceptions.\n\n"
    "The actual RO answer for fair opportunity (shifted to Row 6) fetched:\n"
    "  - FAR_Part_16_IDIQ_Comprehensive_RFO_2025.txt (71,196 chars)\n"
    "  - FAR_Part_16_Contract_Types_RFO_2025.txt (18,174 chars)\n\n"
    "ESSENTIAL SOURCE GAP from the shifted RO answer:\n"
    "  FAR_Part_16_IDIQ_Comprehensive_RFO_2025.txt (71K chars of full Part 16 RFO text)\n"
    "  RO extracted a 9-item J&A content checklist for orders above SAT:\n"
    "    1. Agency/contracting activity identification\n"
    "    2. Nature/description of the action\n"
    "    3. Description of supplies/services + estimated value\n"
    "    4. Identification of exception + supporting rationale\n"
    "    5. Determination that cost is fair and reasonable\n"
    "    6. Supporting facts\n"
    "    7. Statement of actions to remove barriers\n"
    "    8. CO certification (accuracy + completeness)\n"
    "    9. Technical personnel certification\n"
    "  v4 did NOT include this checklist -- it listed the 6 exceptions and approval "
    "tiers but not the required J&A document contents.\n\n"
    "  RO also noted: protest exposure for orders above $10M (civilian agencies) -- v4 missed this.\n\n"
    "VERDICT: v4 missing the 9-item J&A content checklist and $10M protest threshold. "
    "Both are actionable for a CO drafting a J&A for a task order exception."
)

gaps[6] = (
    "MOST SIGNIFICANT GAPS -- RO READ 167K CHARS OF PRIMARY SOURCES\n\n"
    "RO fetched 5 KB documents. Three have material gaps in v4:\n\n"
    "GAP 1 (HIGH): JA_Desk_Guide_January_2025_Updated.txt (57,126 chars)\n"
    "  The current J&A Desk Guide. RO used it for:\n"
    "  - Source-selection-sensitive information handling during debriefings\n"
    "  - What can/cannot be disclosed while a protest is pending\n"
    "  v4 did NOT fetch this document. v4 mentioned the concept but without\n"
    "  the specific current guidance. A CO conducting a debriefing during a\n"
    "  pending protest needs this -- disclosure errors are a common protest ground.\n\n"
    "GAP 2 (MEDIUM): NIH_6315_1_RD_Contracts_Part1_Presolicitation.txt (39,653 chars)\n"
    "  NIH R&D/SBIR-specific policy. RO used it for:\n"
    "  - SBIR dual peer review documentation requirements\n"
    "  - SRG/peer review records must be included in agency report\n"
    "  - BAA vs FAR Part 15 authority distinction for SBIRs\n"
    "  v4 covered BAA caveat conceptually but without NIH-specific procedural details.\n"
    "  Missing peer review documentation requirement is a practical gap -- thin\n"
    "  scientific review records are a common vulnerability in R&D protest reports.\n\n"
    "GAP 3 (LOW-MEDIUM): GAO_Bid_Protests_Essential_Guide_for_COs.txt (17,606 chars)\n"
    "  Full GAO guide. RO used for: 4 CFR 21.2(a)(2) debriefing exception quote,\n"
    "  agency report requirements.\n"
    "  v4 cited GAO-18-510SP but did not fetch/read the file.\n\n"
    "NOT A GAP: NIH_6033_2_Stay_Provisions_Protests.txt (22,462 chars)\n"
    "  Both v4 and RO correctly stated the NIH routing chain and stay procedures.\n"
    "  v4 likely sourced this from model knowledge but got it right.\n\n"
    "NOT A GAP: FAR_Part_15_Negotiation_RFO_2025.txt (10,588 chars)\n"
    "  v4 handled the dual numbering (15.505 vs 15.206-2) better than RO.\n\n"
    "VERDICT: Two essential gaps:\n"
    "  (1) J&A Desk Guide -- debriefing disclosure rules during pending protest\n"
    "  (2) NIH 6315-1 -- SBIR peer review documentation requirements\n"
    "Both are actionable details a CO conducting a debriefing + protest response needs."
)

gaps[7] = (
    "NO SOURCE GAPS (design discussion)\n\n"
    "Neither RO nor v4 cited external sources. This is a product design discussion,\n"
    "not a FAR/regulatory question. Both gave thoughtful analysis from general reasoning.\n\n"
    "VERDICT: No gap. Different but complementary approaches."
)

for row, text in gaps.items():
    ws.cell(row=row, column=col, value=text).alignment = wrap

ws.column_dimensions["O"].width = 100

wb.save(XLSX)
print("Done! Source gap analysis written to column O.\n")

print("=" * 80)
print("SOURCE GAP SUMMARY")
print("=" * 80)

for row in sorted(gaps.keys()):
    q = str(ws.cell(row=row, column=4).value or "")[:70]
    first = gaps[row].split("\n")[0]
    print(f"\nQ{row-1}: {q}")
    print(f"  >> {first}")

print("\n" + "=" * 80)
print("ESSENTIAL SOURCES RO HAD THAT v4 NEEDS")
print("=" * 80)
print("""
Q2 (GAO B-302358):
  - GAO_B-302358_IDIQ_Min_Fund.txt (28K chars) -- full decision text
    v4 got facts right but used a non-verbatim quote

Q4 (Fair Opportunity) [from RO's shifted answer]:
  - FAR_Part_16_IDIQ_Comprehensive_RFO_2025.txt (71K chars)
    v4 missing: 9-item J&A content checklist, $10M protest threshold

Q5 (SBIR Protest):
  - JA_Desk_Guide_January_2025_Updated.txt (57K chars) -- HIGH PRIORITY
    v4 missing: debriefing disclosure rules during pending protest
  - NIH_6315_1_RD_Contracts_Part1_Presolicitation.txt (40K chars) -- MEDIUM
    v4 missing: SBIR peer review documentation requirements

RECOMMENDATION: Enhance the knowledge cascade so that when the agent
encounters protest/debriefing scenarios, it fetches the J&A Desk Guide
and relevant NIH policy manuals BEFORE responding. The compliance matrix
handles thresholds and documents well, but complex procedural scenarios
need deeper KB reads that the agent is currently skipping.
""")
