"""Write V9 baseline scores and comparative judgments to Excel."""
import sys
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

xlsx_path = "Use Case List.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws = wb["Baseline questions"]

# V9 responses in col 44 (AR), scoring cols 45-50 (AS-AX)
scores = {
    2: {
        "acc": 5, "comp": 5, "src": 4, "act": 5, "total": 19,
        "judgment": (
            "EAGLE V9 = RO (comparable)\n\n"
            "RO comparison: Both correctly cite $15K MPT, $350K SAT under FAC 2025-06. "
            "V9 adds MPT exceptions (construction $2K, SCA $2.5K, contingency $25K) and "
            "SAT commercial extensions ($7.5M/$15M). RO adds practical notes about "
            "HHS-653 form and CPARS requirements.\n\n"
            "v8 vs v9: Identical response quality. No change in tool usage or output.\n"
            "Tools: v8=[query_compliance_matrix] | v9=[query_compliance_matrix]\n"
            "Cascade: N/A -- threshold lookup, no KB cascade needed.\n"
            "KB docs cited: V8=0 | V9=0 (unchanged, expected for compliance_matrix queries)\n"
            "Verdict: Stable -- compliance matrix tool delivers accurate threshold data."
        ),
    },
    3: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved)\n\n"
            "RO comparison: Both correctly cover the two core holdings (IDIQ != multiyear, "
            "minimum obligation at award). V9 adds: creating vs recording distinction table, "
            "ADA analysis with no-year vs annual funds scenario, and cites 2 KB document paths.\n\n"
            "v8 vs v9: SOURCE CITATION IMPROVEMENT -- V8 cited 0 KB document paths "
            "(legal_counsel subagent read docs internally but paths were not surfaced). "
            "V9 now cites 2 KB paths: FAR_Part_16_IDIQ_Comprehensive_RFO_2025.txt and "
            "FC_16_Contract_Types_Guidance.txt. This is the direct result of the subagent "
            "KB tracking feature (commit f0d5060).\n"
            "Tools: v8=[knowledge_search, legal_counsel] | v9=[knowledge_search, legal_counsel]\n"
            "Cascade: legal_counsel subagent fetched KB docs in both versions, but V9 now "
            "surfaces the paths in the supervisor response.\n"
            "Source gap: CLOSED -- V8 had 0 KB paths vs RO's 2; V9 now matches with 2 KB paths.\n"
            "KB docs cited: V8=0 | V9=2 (+2 improvement)\n"
            "Verdict: Key improvement -- subagent KB document reads now visible. Sources 4->5."
        ),
    },
    4: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved)\n\n"
            "RO comparison: V9 covers all RO content (severable/non-severable rules, bona fide "
            "needs, GAO legal standard). V9 adds: short bridge exception, multi-year/no-year "
            "appropriations, concrete dollar examples, and 4 KB document paths.\n\n"
            "v8 vs v9: Comparable quality. V9 cites 4 KB paths via research subagent tracking.\n"
            "Tools: v8=[research] | v9=[research]\n"
            "Cascade: Research subagent consistently triggers deep KB fetch.\n"
            "Source gap: CLOSED -- V9 matches or exceeds RO source coverage.\n"
            "KB docs cited: V8=4 | V9=4 (maintained)\n"
            "Verdict: Strong -- maintains V8 quality, research subagent KB tracking stable."
        ),
    },
    5: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved)\n\n"
            "RO comparison: RO column E contains Q4 content (fair opportunity exceptions), "
            "misaligned with Q4 question. V9 correctly answers with all 7 FAR 16.505(b)(2)(i) "
            "exceptions, verbatim statutory text, documentation/approval thresholds, and 1 KB path.\n\n"
            "v8 vs v9: Consistent quality. Both cite FAR_Part_16_IDIQ_Comprehensive_RFO_2025.txt.\n"
            "Tools: v8=[search_far, knowledge_fetch] | v9=[search_far, knowledge_fetch]\n"
            "Cascade: search_far -> knowledge_fetch working correctly.\n"
            "KB docs cited: V8=1 | V9=1 (maintained)\n"
            "Verdict: Stable -- FAR search + fetch cascade reliable for regulatory questions."
        ),
    },
    6: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved)\n\n"
            "RO comparison: Both produce comprehensive SBIR protest analysis. V9 adds: FAR 2.0 "
            "numbering (15.206-2), NIH Manual 6033-2 stay provisions, and 6 KB doc paths. "
            "RO covers similar ground but V9 is more extensively sourced.\n\n"
            "v8 vs v9: V9 cites 6 KB paths. Tools expanded with web_search and web_fetch.\n"
            "Tools: v8=[search_far, research, knowledge_fetch] | "
            "v9=[search_far, research, knowledge_fetch, web_search, web_fetch]\n"
            "Cascade: Full cascade -- FAR + research + KB fetch + web. Maximum depth.\n"
            "Source gap: CLOSED -- V9 covers all RO protest sources plus web references.\n"
            "KB docs cited: V8=6 | V9=6 (maintained)\n"
            "Verdict: Strongest response in suite -- maintains V8 20/20 with additional web sourcing."
        ),
    },
    7: {
        "acc": 5, "comp": 5, "src": 3, "act": 5, "total": 18,
        "judgment": (
            "EAGLE V9 = RO (comparable)\n\n"
            "RO comparison: Both correctly identify constraint-first vs requirements-first "
            "sequencing tension. V9 uses load_skill for intake context. RO provides similar "
            "analysis with 'three walls of the box' framing.\n\n"
            "Tools: v8=N/A (new question) | v9=[load_skill]\n"
            "Cascade: N/A -- design discussion, no regulatory content needed.\n"
            "KB docs cited: V9=0 (expected for design questions)\n"
            "Verdict: Stable -- design reasoning quality consistent."
        ),
    },
    8: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved) [NEW QUESTION]\n\n"
            "RO comparison: Both correctly identify $280K below SAT, requiring SSJ not full "
            "JOFOC. V9 adds complete 15-item document checklist table, sole-source rationale "
            "language, and 4 KB doc paths. RO provides similar analysis with option year warning.\n\n"
            "Tools: v9=[query_compliance_matrix, research]\n"
            "Cascade: Compliance matrix + research subagent -- full cascade.\n"
            "KB docs cited: V9=4 (HHS_PMR_Common_Requirements, File_Reviewers_Checklist_FRC, etc.)\n"
            "Verdict: Strong -- comprehensive sole-source guidance with actionable checklist."
        ),
    },
    9: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved) [NEW QUESTION -- enhanced Q1]\n\n"
            "RO comparison: V9 provides FAR thresholds AND NIH-specific policy layers "
            "(HCA approval tiers, J&A thresholds, purchase card limits). RO covers similar "
            "FAR thresholds plus NIH FAQ. V9 adds HCA/SPE approval threshold table.\n\n"
            "Tools: v9=[query_compliance_matrix, research]\n"
            "Cascade: Compliance matrix + research -- pulls NIH policy overlays.\n"
            "KB docs cited: V9=4 (HHS_PMR_SAP_Checklist, Common_Requirements, FRC, NIH_Policy_6307)\n"
            "Verdict: Enhanced Q1 answered comprehensively -- NIH layers add value."
        ),
    },
    10: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved) [NEW QUESTION -- enhanced Q2]\n\n"
            "RO comparison: V9 correctly identifies B-302358 not present as standalone KB "
            "file, provides transparent sourcing note, draws from FAR Part 16 IDIQ and web. "
            "RO had direct KB file access. Both cover core holdings correctly.\n\n"
            "Tools: v9=[knowledge_search, knowledge_fetch, web_search, web_fetch]\n"
            "Cascade: Full cascade -- KB search + fetch + web supplementation.\n"
            "KB docs cited: V9=3 (FAR_Part_16_IDIQ, agents/02-legal, + web sources)\n"
            "Verdict: Strong -- transparent about source limitations, supplements with web."
        ),
    },
    11: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved) [NEW QUESTION -- enhanced Q3]\n\n"
            "RO comparison: V9 provides the most comprehensive appropriations law response "
            "in the suite (12,595 chars). Covers statutory foundation, performance period "
            "rule, severable/non-severable, 10 USC 2410a, option exercise, edge cases.\n\n"
            "Tools: v9=[search_far, research]\n"
            "Cascade: FAR search + research subagent -- deep appropriations analysis.\n"
            "KB docs cited: V9=5 (severable, bona_fide_needs, options, IDIQ_funding, GAO_Guidance)\n"
            "Verdict: Exceptional -- most thorough appropriations response with edge cases."
        ),
    },
    12: {
        "acc": 5, "comp": 5, "src": 4, "act": 5, "total": 19,
        "judgment": (
            "EAGLE V9 = RO (comparable) [NEW QUESTION -- enhanced Q4]\n\n"
            "RO comparison: Both list all 7 FAR 16.505(b)(2)(i) exceptions with verbatim "
            "text. V9 includes exact paragraph citations and approval dollar thresholds. "
            "RO had web-supplemented verification.\n\n"
            "Tools: v9=[search_far, knowledge_fetch]\n"
            "Cascade: search_far -> knowledge_fetch -- standard regulatory lookup.\n"
            "KB docs cited: V9=1 (FAR_Part_16_IDIQ_Comprehensive_RFO_2025)\n"
            "Verdict: Strong regulatory response -- verbatim text adds value. Sources adequate."
        ),
    },
    13: {
        "acc": 5, "comp": 5, "src": 5, "act": 5, "total": 20,
        "judgment": (
            "EAGLE V9 > RO (improved) [NEW QUESTION -- enhanced Q5/SBIR protest]\n\n"
            "RO comparison: V9 provides KB document inventory FIRST (11 documents with full "
            "S3 paths), then substantive protest analysis. RO pulls similar docs but V9's "
            "format is more transparent and reproducible.\n\n"
            "Tools: v9=[knowledge_search, knowledge_fetch]\n"
            "Cascade: KB search -> fetch -- deep retrieval (11 unique KB paths). Highest "
            "source count in entire baseline suite.\n"
            "KB docs cited: V9=11 (GAO Essential Guide, FC_33, NIH 6033-2, NIH 6033-1, etc.)\n"
            "Verdict: Exceptional -- transparent KB inventory + comprehensive analysis."
        ),
    },
    14: {
        "acc": 4, "comp": 3, "src": 4, "act": 3, "total": 14,
        "judgment": (
            "EAGLE V9 < RO (regression) [NEW QUESTION -- SOW generation]\n\n"
            "RO comparison: RO provides 1,907 chars of context about the generated SOW "
            "including key sections, scope elements, and contract details. V9 returns only "
            "255 chars saying 'Generated a draft sow document.' The actual SOW exists as a "
            "document artifact, but the chat response lacks summary or preview.\n\n"
            "Tools: v9=[create_document]\n"
            "Cascade: Document generation uses correct template (SOW Template Eagle v2).\n"
            "KB docs cited: V9=1 (template path only)\n"
            "Verdict: Document generated correctly but chat response too terse. User must "
            "navigate to document card. create_document should echo key sections back."
        ),
    },
}

wrap = Alignment(wrap_text=True, vertical="top")
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Write headers (cols 45-50)
headers = [
    (45, "EAGLE V9 Accuracy (0-5)"),
    (46, "EAGLE V9 Completeness (0-5)"),
    (47, "EAGLE V9 Sources (0-5)"),
    (48, "EAGLE V9 Actionability (0-5)"),
    (49, "EAGLE V9 Total (0-20)"),
    (50, "V9 vs RO + V8 Comparative Judgment"),
]

for col, text in headers:
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

# Write scores
for row, s in scores.items():
    ws.cell(row=row, column=45, value=s["acc"]).alignment = wrap
    ws.cell(row=row, column=46, value=s["comp"]).alignment = wrap
    ws.cell(row=row, column=47, value=s["src"]).alignment = wrap
    ws.cell(row=row, column=48, value=s["act"]).alignment = wrap
    ws.cell(row=row, column=49, value=s["total"]).alignment = wrap
    ws.cell(row=row, column=50, value=s["judgment"]).alignment = wrap

# Set column widths
for col in range(45, 50):
    letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[letter].width = 14
ws.column_dimensions[openpyxl.utils.get_column_letter(50)].width = 100

wb.save(xlsx_path)
print("Scores written to Excel successfully.")

# Save scores JSON
scores_json = {}
for row, s in scores.items():
    q_num = row - 1
    scores_json[str(q_num)] = {
        "acc": s["acc"],
        "comp": s["comp"],
        "src": s["src"],
        "act": s["act"],
        "total": s["total"],
        "verdict": s["judgment"].split("\n")[0],
    }
with open("scripts/baseline_v9_scores.json", "w", encoding="utf-8") as f:
    json.dump(scores_json, f, indent=2, ensure_ascii=False)
print("Scores JSON saved to scripts/baseline_v9_scores.json")
