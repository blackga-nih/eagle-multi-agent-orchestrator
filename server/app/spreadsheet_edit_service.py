"""XLSX preview extraction and structured editing helpers."""

from __future__ import annotations

import io
import logging
import os
import re
from dataclasses import dataclass
from typing import Any, Optional

from botocore.exceptions import BotoCoreError, ClientError

from .changelog_store import write_document_changelog_entry
from .db_client import get_s3
from .document_key_utils import (
    extract_package_document_ref,
    extract_workspace_document_ref,
    is_allowed_document_key,
)
from .document_service import create_package_document_version
from .document_store import get_document
from .formula_evaluation import evaluate_workbook_formulas, evaluate_workbook_formulas_safe
from .template_service import XLSXPopulator

logger = logging.getLogger("eagle.spreadsheet_edit")

S3_BUCKET = os.getenv("S3_BUCKET", "eagle-documents-695681773636-dev")
MAX_PREVIEW_ROWS = 80
MAX_PREVIEW_COLS = 18


@dataclass
class SpreadsheetCellEdit:
    sheet_id: str
    cell_ref: str
    value: str


def _serialize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _coerce_cell_value(value: str) -> Any:
    stripped = value.strip()
    if stripped == "":
        return None
    if re.fullmatch(r"[-+]?\d+", stripped):
        try:
            return int(stripped)
        except ValueError:
            return value
    if re.fullmatch(r"[-+]?\d*\.\d+", stripped):
        try:
            return float(stripped)
        except ValueError:
            return value
    lowered = stripped.lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    return value


def _sheet_identity(ws) -> str:
    return re.sub(r"[^a-z0-9]+", "-", ws.title.lower()).strip("-") or "sheet"


def extract_xlsx_preview_payload(xlsx_bytes: bytes) -> dict[str, Any]:
    """Return text plus structured worksheets for browser preview/editing."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "content": "[XLSX preview unavailable - openpyxl not installed]",
            "preview_mode": "none",
            "preview_sheets": [],
        }

    # Evaluate formulas first so data_only=True gets calculated values
    evaluated_bytes = evaluate_workbook_formulas_safe(xlsx_bytes)

    wb = load_workbook(io.BytesIO(evaluated_bytes), data_only=False)
    wb_values = load_workbook(io.BytesIO(evaluated_bytes), data_only=True)
    preview_sheets: list[dict[str, Any]] = []

    for sheet_index, ws in enumerate(wb.worksheets):
        if ws.sheet_state != "visible":
            continue

        values_ws = wb_values[ws.title]
        row_limit = min(ws.max_row or 1, MAX_PREVIEW_ROWS)
        col_limit = min(ws.max_column or 1, MAX_PREVIEW_COLS)
        hidden_cells: set[str] = set()
        for merged_range in ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_ref = ws.cell(row=row, column=col).coordinate
                    if row != merged_range.min_row or col != merged_range.min_col:
                        hidden_cells.add(cell_ref)

        rows: list[dict[str, Any]] = []
        for row_idx in range(1, row_limit + 1):
            cells: list[dict[str, Any]] = []
            row_has_content = False
            for col_idx in range(1, col_limit + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value_cell = values_ws.cell(row=row_idx, column=col_idx)
                raw_value = cell.value
                is_formula = isinstance(raw_value, str) and raw_value.startswith("=")
                # Determine display value: prefer cached calculation, avoid showing raw formulas
                cached_value = value_cell.value
                if is_formula:
                    # For formula cells: use cached value if it's NOT the formula string
                    # (data_only=True should give calculated value, but double-check)
                    if cached_value is not None and not (isinstance(cached_value, str) and cached_value.startswith("=")):
                        display_value = cached_value
                    else:
                        # No valid cached value - show empty rather than formula text
                        display_value = ""
                elif cached_value is not None:
                    display_value = cached_value
                else:
                    display_value = raw_value
                is_hidden = cell.coordinate in hidden_cells
                editable = not is_formula and not is_hidden
                cell_payload = {
                    "cell_ref": cell.coordinate,
                    "row": row_idx,
                    "col": col_idx,
                    "value": _serialize_cell_value(raw_value),
                    "display_value": _serialize_cell_value(display_value),
                    "editable": editable,
                    "is_formula": is_formula,
                }
                if cell_payload["display_value"]:
                    row_has_content = True
                cells.append(cell_payload)

            if row_has_content or any(cell["editable"] for cell in cells):
                rows.append({"row_index": row_idx, "cells": cells})

        if rows:
            preview_sheets.append(
                {
                    "sheet_id": f"{sheet_index}:{_sheet_identity(ws)}",
                    "title": ws.title,
                    "max_row": row_limit,
                    "max_col": col_limit,
                    "truncated": ws.max_row > row_limit or ws.max_column > col_limit,
                    "rows": rows,
                }
            )

    return {
        "content": XLSXPopulator.extract_text(xlsx_bytes),
        "preview_mode": "xlsx_grid",
        "preview_sheets": preview_sheets,
    }


def apply_xlsx_cell_edits(
    xlsx_bytes: bytes, cell_edits: list[SpreadsheetCellEdit]
) -> tuple[bytes, int, list[str]]:
    """Apply structured cell edits back to a workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    sheet_map = {
        f"{idx}:{_sheet_identity(ws)}": ws
        for idx, ws in enumerate(wb.worksheets)
        if ws.sheet_state == "visible"
    }
    applied = 0
    missing: list[str] = []

    for edit in cell_edits:
        ws = sheet_map.get(edit.sheet_id)
        if ws is None:
            missing.append(f"{edit.sheet_id}:{edit.cell_ref}")
            continue
        cell = ws[edit.cell_ref]
        raw_value = cell.value
        if isinstance(raw_value, str) and raw_value.startswith("="):
            missing.append(f"{edit.sheet_id}:{edit.cell_ref}")
            continue
        next_value = _coerce_cell_value(edit.value)
        if cell.value != next_value:
            cell.value = next_value
            applied += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), applied, missing


def save_xlsx_preview_edits(
    *,
    tenant_id: str,
    user_id: Optional[str],
    doc_key: str,
    cell_edits: list[dict[str, Any]],
    session_id: Optional[str] = None,
    change_source: str = "user_edit",
) -> dict[str, Any]:
    """Persist browser-side XLSX cell edits."""
    logger.info("[XLSX Save Debug] Received cell_edits=%s, doc_key=%s", cell_edits, doc_key)
    if not doc_key:
        return {"error": "document_key is required"}
    if not cell_edits:
        logger.warning("[XLSX Save Debug] No cell_edits provided!")
        return {"error": "cell_edits are required"}
    if not is_allowed_document_key(doc_key, tenant_id, user_id):
        return {"error": "Access denied for document key"}
    if not doc_key.lower().endswith(".xlsx"):
        return {"error": "Structured spreadsheet editing only supports .xlsx documents"}

    edits = [
        SpreadsheetCellEdit(
            sheet_id=str(edit.get("sheet_id", "")),
            cell_ref=str(edit.get("cell_ref", "")),
            value=str(edit.get("value", "")),
        )
        for edit in cell_edits
        if edit.get("sheet_id") and edit.get("cell_ref") is not None
    ]
    if not edits:
        return {"error": "No valid spreadsheet cell edits were provided"}

    s3 = get_s3()
    try:
        response = s3.get_object(Bucket=S3_BUCKET, Key=doc_key)
        original_bytes = response["Body"].read()
        content_type = (
            response.get("ContentType")
            or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except (ClientError, BotoCoreError) as exc:
        logger.error("Failed to load XLSX preview artifact: %s", exc, exc_info=True)
        return {"error": "Failed to load document."}

    try:
        updated_bytes, applied_count, missing = apply_xlsx_cell_edits(
            original_bytes, edits
        )
        logger.info("[XLSX Save Debug] Applied %d edits, missing=%s", applied_count, missing)
    except Exception as exc:
        logger.error("Failed to apply structured XLSX edits: %s", exc, exc_info=True)
        return {"error": "Failed to apply spreadsheet edits."}

    if applied_count == 0:
        # No actual changes needed - return success with current state (not an error)
        # This happens when user saves the same value that's already in the cell
        preview_payload = extract_xlsx_preview_payload(original_bytes)
        return {
            "success": True,
            "mode": "no_changes",
            "message": "No changes were needed.",
            "content": preview_payload.get("content"),
            "preview_mode": preview_payload.get("preview_mode"),
            "preview_sheets": preview_payload.get("preview_sheets", []),
            "missing": missing,
        }

    # Evaluate formulas for PREVIEW only - don't save evaluated bytes (preserves formulas)
    preview_bytes, formulas_evaluated = evaluate_workbook_formulas(updated_bytes)
    preview_payload = extract_xlsx_preview_payload(preview_bytes)

    package_ref = extract_package_document_ref(doc_key)
    if package_ref:
        if package_ref["tenant_id"] != tenant_id:
            return {"error": "Access denied for package document"}
        package_id = str(package_ref["package_id"])
        doc_type = str(package_ref["doc_type"])
        version = int(package_ref["version"])
        existing = get_document(tenant_id, package_id, doc_type, version)
        title = (existing or {}).get("title") or doc_type.replace("_", " ").title()
        # Save original bytes WITH formulas intact (not the evaluated preview_bytes)
        result = create_package_document_version(
            tenant_id=tenant_id,
            package_id=package_id,
            doc_type=doc_type,
            content=updated_bytes,
            title=title,
            file_type="xlsx",
            created_by_user_id=user_id,
            session_id=session_id,
            change_source=change_source,
            template_id=(existing or {}).get("template_id"),
        )
        if not result.success:
            return {"error": result.error or "Failed to save document version"}
        message = f"Saved spreadsheet version {result.version}."
        if not formulas_evaluated:
            message += " Note: Formulas will calculate when opened in Excel."
        # Generate presigned download URL for the saved document
        download_url = None
        try:
            download_url = s3.generate_presigned_url(
                "get_object",
                Params={"Bucket": S3_BUCKET, "Key": result.s3_key},
                ExpiresIn=3600,
            )
        except Exception as exc:
            logger.warning("Failed to generate presigned URL: %s", exc)
        return {
            "success": True,
            "mode": "package_xlsx_preview_edit",
            "document_id": result.document_id,
            "key": result.s3_key,
            "version": result.version,
            "file_type": "xlsx",
            "content": preview_payload.get("content"),
            "preview_mode": preview_payload.get("preview_mode"),
            "preview_sheets": preview_payload.get("preview_sheets", []),
            "missing": missing,
            "formulas_calculated": formulas_evaluated,
            "message": message,
            "download_url": download_url,
        }

    workspace_ref = extract_workspace_document_ref(doc_key)
    if not workspace_ref:
        return {"error": "Unsupported XLSX key format"}
    if workspace_ref["tenant"] != tenant_id or workspace_ref["user"] != user_id:
        return {"error": "Access denied for workspace document"}

    try:
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=doc_key,
            Body=updated_bytes,
            ContentType=content_type,
        )
        write_document_changelog_entry(
            tenant_id=tenant_id,
            document_key=doc_key,
            change_type="update",
            change_source=change_source,
            change_summary="Updated XLSX via preview editor",
            actor_user_id=user_id or "user",
        )
    except Exception as exc:
        logger.error(
            "Failed to save workspace XLSX preview edits: %s", exc, exc_info=True
        )
        return {"error": "Failed to save spreadsheet."}

    ws_message = "Spreadsheet saved."
    if not formulas_evaluated:
        ws_message += " Note: Formulas will calculate when opened in Excel."
    # Generate presigned download URL for the saved document
    ws_download_url = None
    try:
        ws_download_url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": S3_BUCKET, "Key": doc_key},
            ExpiresIn=3600,
        )
    except Exception as exc:
        logger.warning("Failed to generate presigned URL: %s", exc)
    return {
        "success": True,
        "mode": "workspace_xlsx_preview_edit",
        "document_id": doc_key,
        "key": doc_key,
        "version": 0,
        "file_type": "xlsx",
        "content": preview_payload.get("content"),
        "preview_mode": preview_payload.get("preview_mode"),
        "preview_sheets": preview_payload.get("preview_sheets", []),
        "missing": missing,
        "formulas_calculated": formulas_evaluated,
        "message": ws_message,
        "download_url": ws_download_url,
    }
