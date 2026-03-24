"""Document Markdown Service — Convert binary documents to markdown text.

Supports PDF, DOCX, XLSX, and plain text/markdown passthrough.
Used during upload to persist a parsed markdown equivalent alongside the original.
"""

from __future__ import annotations

import io
import logging
from typing import Optional

logger = logging.getLogger("eagle.document_markdown")


def convert_to_markdown(
    body: bytes,
    content_type: str,
    filename: str = "",
) -> Optional[str]:
    """Convert uploaded document bytes to markdown text.

    Args:
        body: Raw file bytes
        content_type: MIME content type
        filename: Original filename (used for fallback type detection)

    Returns:
        Markdown string or None if conversion is not supported/fails.
    """
    if not body:
        return None

    try:
        # Plain text / Markdown — passthrough
        if content_type in ("text/plain", "text/markdown"):
            return _plaintext_to_markdown(body)

        # PDF
        if content_type == "application/pdf":
            return _pdf_to_markdown(body)

        # DOCX
        if content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return _docx_to_markdown(body)

        # Legacy DOC
        if content_type == "application/msword":
            # Best-effort: try python-docx (sometimes works with .doc)
            try:
                return _docx_to_markdown(body)
            except Exception:
                logger.debug("Legacy .doc conversion failed, skipping")
                return None

        # XLSX / XLS
        if content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return _xlsx_to_markdown(body)

        # Fallback: check filename extension
        lower_name = (filename or "").lower()
        if lower_name.endswith(".md") or lower_name.endswith(".txt"):
            return _plaintext_to_markdown(body)
        if lower_name.endswith(".pdf"):
            return _pdf_to_markdown(body)
        if lower_name.endswith(".docx"):
            return _docx_to_markdown(body)
        if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            return _xlsx_to_markdown(body)

        return None

    except Exception as e:
        logger.warning("Markdown conversion failed for %s: %s", content_type, e)
        return None


def _plaintext_to_markdown(body: bytes) -> str:
    """Passthrough for plain text and markdown files."""
    return body.decode("utf-8", errors="replace")


def _pdf_to_markdown(body: bytes) -> Optional[str]:
    """Extract text from PDF and format as markdown with page breaks."""
    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning("pypdf not installed, skipping PDF conversion")
        return None

    reader = PdfReader(io.BytesIO(body))
    parts: list[str] = []

    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            if i > 0:
                parts.append(f"\n---\n\n<!-- Page {i + 1} -->\n")
            parts.append(text)

    return "\n".join(parts) if parts else None


def _docx_to_markdown(body: bytes) -> Optional[str]:
    """Convert DOCX to markdown preserving headings and tables."""
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx not installed, skipping DOCX conversion")
        return None

    doc = Document(io.BytesIO(body))
    parts: list[str] = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

        if tag == "p":
            _convert_paragraph(element, doc, parts)
        elif tag == "tbl":
            _convert_table(element, doc, parts)

    return "\n".join(parts) if parts else None


def _convert_paragraph(element, doc, parts: list[str]) -> None:
    """Convert a DOCX paragraph element to markdown."""
    from docx.oxml.ns import qn

    # Get paragraph style
    style_elem = element.find(qn("w:pPr"))
    style_name = ""
    if style_elem is not None:
        p_style = style_elem.find(qn("w:pStyle"))
        if p_style is not None:
            style_name = p_style.get(qn("w:val"), "")

    # Extract text
    text = "".join(
        node.text or ""
        for node in element.iter(qn("w:t"))
    )

    if not text.strip():
        if parts and parts[-1] != "":
            parts.append("")
        return

    # Convert heading styles
    heading_level = 0
    if style_name.startswith("Heading"):
        try:
            heading_level = int(style_name.replace("Heading", "").strip())
        except ValueError:
            heading_level = 0

    # Check for outline level in paragraph properties
    if heading_level == 0 and style_elem is not None:
        outline = style_elem.find(qn("w:outlineLvl"))
        if outline is not None:
            try:
                heading_level = int(outline.get(qn("w:val"), "0")) + 1
            except ValueError:
                pass

    if heading_level > 0:
        prefix = "#" * min(heading_level, 6)
        parts.append(f"\n{prefix} {text.strip()}\n")
    else:
        # Check for list items
        num_pr = None
        if style_elem is not None:
            num_pr = style_elem.find(qn("w:numPr"))

        if num_pr is not None:
            parts.append(f"- {text.strip()}")
        else:
            parts.append(text.strip())


def _convert_table(element, doc, parts: list[str]) -> None:
    """Convert a DOCX table element to a markdown pipe table."""
    from docx.oxml.ns import qn

    rows: list[list[str]] = []

    for tr in element.iter(qn("w:tr")):
        cells: list[str] = []
        for tc in tr.iter(qn("w:tc")):
            cell_text = " ".join(
                node.text or ""
                for node in tc.iter(qn("w:t"))
            ).strip()
            cells.append(cell_text)
        if cells:
            rows.append(cells)

    if not rows:
        return

    parts.append("")  # blank line before table

    # Normalize column count
    max_cols = max(len(r) for r in rows)
    for row in rows:
        while len(row) < max_cols:
            row.append("")

    # Header row
    header = rows[0]
    parts.append("| " + " | ".join(header) + " |")
    parts.append("| " + " | ".join("---" for _ in header) + " |")

    # Data rows
    for row in rows[1:]:
        parts.append("| " + " | ".join(row) + " |")

    parts.append("")  # blank line after table


def _xlsx_to_markdown(body: bytes) -> Optional[str]:
    """Convert XLSX spreadsheet to markdown tables, one per visible sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed, skipping XLSX conversion")
        return None

    wb = load_workbook(io.BytesIO(body), data_only=True, read_only=True)
    parts: list[str] = []

    for ws in wb.worksheets:
        if hasattr(ws, "sheet_state") and ws.sheet_state != "visible":
            continue

        sheet_rows: list[list[str]] = []
        for row in ws.iter_rows(max_row=100, values_only=True):
            cell_values = [str(c) if c is not None else "" for c in row]
            if any(v for v in cell_values):
                sheet_rows.append(cell_values)

        if not sheet_rows:
            continue

        parts.append(f"\n## Sheet: {ws.title}\n")

        # Normalize column count
        max_cols = max(len(r) for r in sheet_rows)
        for row in sheet_rows:
            while len(row) < max_cols:
                row.append("")

        # First row as header
        header = sheet_rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")

        for row in sheet_rows[1:]:
            parts.append("| " + " | ".join(row) + " |")

        parts.append("")

    wb.close()
    return "\n".join(parts) if parts else None
