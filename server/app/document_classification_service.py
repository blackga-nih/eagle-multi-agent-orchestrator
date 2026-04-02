"""
Document Classification Service — Classify uploaded documents by type.

Provides filename-based heuristics with optional content-based AI fallback
for ambiguous cases. Supports text extraction from PDF, DOCX, XLSX, and plain text.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from typing import Optional

from .doc_type_registry import ALL_DOC_TYPES

logger = logging.getLogger("eagle.document_classification")

# Valid document types — imported from the centralized registry
VALID_DOC_TYPES = ALL_DOC_TYPES

# Filename patterns for quick classification (most specific first).
# Original 5 core types kept at top for fast-path matching.
FILENAME_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    # ── Core 5 (highest traffic) ──────────────────────────────────────
    # Statement of Work
    (re.compile(r"\b(sow|statement[-_\s]?of[-_\s]?work)\b", re.IGNORECASE), "sow"),
    # IGCE / Cost Estimate
    (
        re.compile(
            r"\b(igce|ige|independent[-_\s]?government[-_\s]?(?:cost[-_\s]?)?estimate|cost[-_\s]?estimate)\b",
            re.IGNORECASE,
        ),
        "igce",
    ),
    # Market Research — use lookahead instead of \b at end so "Market_Research_Report" matches
    # (\b fails when next char is _ because _ is a word character)
    (re.compile(r"\b(market[-_\s]?research|mrr)(?:\b|(?=[-_\s.]))", re.IGNORECASE), "market_research"),
    # Justification
    (
        re.compile(
            r"\b(j[&]?a|justification(?:[-_\s]?(?:&|and)[-_\s]?approval)?|sole[-_\s]?source)\b",
            re.IGNORECASE,
        ),
        "justification",
    ),
    # Acquisition Plan
    (
        re.compile(
            r"\b(acquisition[-_\s]?plan|streamlined[-_\s]?acquisition[-_\s]?plan)\b",
            re.IGNORECASE,
        ),
        "acquisition_plan",
    ),
    # ── Extended types (17 additional categories) ─────────────────────
    # NOTE: Use (?:^|[\W_]) instead of \b for patterns that appear in
    # filenames with underscore separators, since Python's \b treats _ as
    # a word character and won't match at _X boundaries.
    # Statement of Need — Products (must come before generic SON match)
    (re.compile(r"(?:^|[\W_])son[-_\s.].*product", re.IGNORECASE), "son_products"),
    (
        re.compile(r"statement[-_\s]?of[-_\s]?need[-_\s.].*product", re.IGNORECASE),
        "son_products",
    ),
    # Statement of Need — Services
    (re.compile(r"(?:^|[\W_])son[-_\s.].*service", re.IGNORECASE), "son_services"),
    (
        re.compile(r"statement[-_\s]?of[-_\s]?need[-_\s.].*service", re.IGNORECASE),
        "son_services",
    ),
    # COR Certification / Appointment
    (
        re.compile(
            r"(?:^|[\W_])cor[-_\s]?(?:appointment|certification|memorandum)",
            re.IGNORECASE,
        ),
        "cor_certification",
    ),
    (re.compile(r"appointment[-_\s]?memorandum", re.IGNORECASE), "cor_certification"),
    # Buy American Act
    (re.compile(r"buy[-_\s]?american", re.IGNORECASE), "buy_american"),
    # Subcontracting Plan
    (re.compile(r"sub[-_\s]?k[-_\s.].*plan", re.IGNORECASE), "subk_plan"),
    (re.compile(r"subcontracting[-_\s]?plan", re.IGNORECASE), "subk_plan"),
    # Subcontracting Review
    (re.compile(r"sub[-_\s]?k[-_\s.].*review", re.IGNORECASE), "subk_review"),
    (re.compile(r"subcontracting[-_\s]?review", re.IGNORECASE), "subk_review"),
    # Conference Waiver (must precede conference_request — "Request for Waiver" matches both)
    (re.compile(r"conference[-_\s.].*waiver", re.IGNORECASE), "conference_waiver"),
    # Conference Request
    (re.compile(r"conference[-_\s.].*request", re.IGNORECASE), "conference_request"),
    (
        re.compile(r"conference[-_\s.].*grant[-_\s]?request", re.IGNORECASE),
        "conference_request",
    ),
    (re.compile(r"conference[-_\s.].*approval", re.IGNORECASE), "conference_request"),
    # Promotional Item
    (re.compile(r"promotional[-_\s]?item", re.IGNORECASE), "promotional_item"),
    # Exemption Determination
    (
        re.compile(r"exemption[-_\s]?determination", re.IGNORECASE),
        "exemption_determination",
    ),
    # Mandatory Use Waiver
    (
        re.compile(r"mandatory[-_\s]?use[-_\s]?waiver", re.IGNORECASE),
        "mandatory_use_waiver",
    ),
    # GFP Form
    (re.compile(r"(?:^|[\W_])gfp[-_\s]?form", re.IGNORECASE), "gfp_form"),
    (
        re.compile(r"government[-_\s]?furnished[-_\s]?property", re.IGNORECASE),
        "gfp_form",
    ),
    # BPA Call Order
    (
        re.compile(r"(?:^|[\W_])bpa[-_\s]?call[-_\s]?order", re.IGNORECASE),
        "bpa_call_order",
    ),
    (
        re.compile(r"blanket[-_\s]?purchase[-_\s]?agreement", re.IGNORECASE),
        "bpa_call_order",
    ),
    (re.compile(r"(?:^|[\W_])bpa[-_\s]?callorder", re.IGNORECASE), "bpa_call_order"),
    # Technical Questionnaire (handles common misspelling "questionnare")
    (
        re.compile(r"technical[-_\s]?questionnai?r?e", re.IGNORECASE),
        "technical_questionnaire",
    ),
    (
        re.compile(r"project[-_\s]?officer.*questionnai?r?e", re.IGNORECASE),
        "technical_questionnaire",
    ),
    # Quotation Abstract
    (re.compile(r"quotation[-_\s]?abstract", re.IGNORECASE), "quotation_abstract"),
    # Receiving Report
    (re.compile(r"receiving[-_\s]?report", re.IGNORECASE), "receiving_report"),
    # SRB Request
    (re.compile(r"(?:^|[\W_])srb[-_\s]?request", re.IGNORECASE), "srb_request"),
    (re.compile(r"source[-_\s]?review[-_\s]?board", re.IGNORECASE), "srb_request"),
    # Reference Guide (lowest priority — generic terms)
    (re.compile(r"structure[-_\s]?guide", re.IGNORECASE), "reference_guide"),
]

# Content patterns for AI fallback (keywords found in document body)
CONTENT_PATTERNS: list[tuple[re.Pattern[str], str, float]] = [
    # ── Core 5 ────────────────────────────────────────────────────────
    # SOW indicators
    (re.compile(r"\bstatement\s+of\s+work\b", re.IGNORECASE), "sow", 0.9),
    (re.compile(r"\bperiod\s+of\s+performance\b", re.IGNORECASE), "sow", 0.7),
    (
        re.compile(r"\bdeliverables?\b.*\btasks?\b", re.IGNORECASE | re.DOTALL),
        "sow",
        0.6,
    ),
    (re.compile(r"\bscope\s+of\s+work\b", re.IGNORECASE), "sow", 0.8),
    # IGCE indicators
    (
        re.compile(
            r"\bindependent\s+government\s+(?:cost\s+)?estimate\b", re.IGNORECASE
        ),
        "igce",
        0.95,
    ),
    (re.compile(r"\bigce\b", re.IGNORECASE), "igce", 0.9),
    (
        re.compile(
            r"\blabor\s+(?:hours?|rates?)\b.*\bcost\b", re.IGNORECASE | re.DOTALL
        ),
        "igce",
        0.7,
    ),
    (
        re.compile(r"\btotal\s+estimated\s+(?:cost|price)\b", re.IGNORECASE),
        "igce",
        0.75,
    ),
    # Market Research indicators
    (re.compile(r"\bmarket\s+research\b", re.IGNORECASE), "market_research", 0.9),
    (re.compile(r"\bvendor\s+analysis\b", re.IGNORECASE), "market_research", 0.8),
    (re.compile(r"\bcompetitive\s+landscape\b", re.IGNORECASE), "market_research", 0.7),
    # Justification indicators
    (
        re.compile(r"\bjustification\s+(?:&|and)\s+approval\b", re.IGNORECASE),
        "justification",
        0.95,
    ),
    (
        re.compile(r"\bsole\s+source\s+justification\b", re.IGNORECASE),
        "justification",
        0.9,
    ),
    (re.compile(r"\bfar\s+6\.3", re.IGNORECASE), "justification", 0.8),
    # Acquisition Plan indicators
    (re.compile(r"\bacquisition\s+plan\b", re.IGNORECASE), "acquisition_plan", 0.9),
    (
        re.compile(r"\bacquisition\s+strategy\b", re.IGNORECASE),
        "acquisition_plan",
        0.85,
    ),
    (
        re.compile(r"\bcontract\s+type\s+selection\b", re.IGNORECASE),
        "acquisition_plan",
        0.75,
    ),
    # ── Extended types ────────────────────────────────────────────────
    # SON Products
    (
        re.compile(r"\bstatement\s+of\s+need\b.*\bproduct", re.IGNORECASE | re.DOTALL),
        "son_products",
        0.9,
    ),
    (re.compile(r"\bequipment\s+and\s+supplies\b", re.IGNORECASE), "son_products", 0.8),
    # SON Services
    (
        re.compile(r"\bstatement\s+of\s+need\b.*\bservice", re.IGNORECASE | re.DOTALL),
        "son_services",
        0.9,
    ),
    (re.compile(r"\bcatalog\s+pricing\b", re.IGNORECASE), "son_services", 0.75),
    # COR Certification
    (re.compile(r"\bcor\s+appointment\b", re.IGNORECASE), "cor_certification", 0.95),
    (
        re.compile(r"\bcontracting\s+officer\s+representative\b", re.IGNORECASE),
        "cor_certification",
        0.9,
    ),
    (re.compile(r"\bfac[-\s]?cor\s+level\b", re.IGNORECASE), "cor_certification", 0.85),
    # Buy American
    (re.compile(r"\bbuy\s+american\b", re.IGNORECASE), "buy_american", 0.9),
    (
        re.compile(r"\bnon[-\s]?availability\s+determination\b", re.IGNORECASE),
        "buy_american",
        0.85,
    ),
    # Subcontracting Plan
    (re.compile(r"\bsubcontracting\s+plan\b", re.IGNORECASE), "subk_plan", 0.9),
    (
        re.compile(r"\bsmall\s+business\s+subcontracting\b", re.IGNORECASE),
        "subk_plan",
        0.85,
    ),
    # Subcontracting Review
    (re.compile(r"\bsubcontracting\s+review\b", re.IGNORECASE), "subk_review", 0.9),
    # Conference Request
    (re.compile(r"\bconference\s+request\b", re.IGNORECASE), "conference_request", 0.9),
    (re.compile(r"\bnih\s+conference\b", re.IGNORECASE), "conference_request", 0.8),
    # Conference Waiver
    (re.compile(r"\bconference\s+waiver\b", re.IGNORECASE), "conference_waiver", 0.9),
    (
        re.compile(
            r"\brequest\s+for\s+waiver\b.*\bconference\b", re.IGNORECASE | re.DOTALL
        ),
        "conference_waiver",
        0.85,
    ),
    # Promotional Item
    (re.compile(r"\bpromotional\s+item\b", re.IGNORECASE), "promotional_item", 0.9),
    # Exemption Determination
    (
        re.compile(r"\bexemption\s+determination\b", re.IGNORECASE),
        "exemption_determination",
        0.9,
    ),
    # Mandatory Use Waiver
    (
        re.compile(r"\bmandatory[-\s]?use\s+waiver\b", re.IGNORECASE),
        "mandatory_use_waiver",
        0.9,
    ),
    # GFP Form
    (
        re.compile(r"\bgovernment\s+furnished\s+property\b", re.IGNORECASE),
        "gfp_form",
        0.9,
    ),
    (re.compile(r"\bgfp\b", re.IGNORECASE), "gfp_form", 0.7),
    # BPA Call Order
    (re.compile(r"\bbpa\s+call\s+order\b", re.IGNORECASE), "bpa_call_order", 0.9),
    (
        re.compile(r"\bblanket\s+purchase\s+agreement\b", re.IGNORECASE),
        "bpa_call_order",
        0.85,
    ),
    # Technical Questionnaire
    (
        re.compile(r"\btechnical\s+questionnai?r?e\b", re.IGNORECASE),
        "technical_questionnaire",
        0.9,
    ),
    (
        re.compile(r"\bproject\s+officer\b", re.IGNORECASE),
        "technical_questionnaire",
        0.6,
    ),
    # Quotation Abstract
    (re.compile(r"\bquotation\s+abstract\b", re.IGNORECASE), "quotation_abstract", 0.9),
    # Receiving Report
    (re.compile(r"\breceiving\s+report\b", re.IGNORECASE), "receiving_report", 0.9),
    # SRB Request
    (re.compile(r"\bsrb\s+request\b", re.IGNORECASE), "srb_request", 0.9),
    (re.compile(r"\bsource\s+review\s+board\b", re.IGNORECASE), "srb_request", 0.85),
]


@dataclass
class ClassificationResult:
    """Result of document classification."""

    doc_type: str
    confidence: float
    method: str  # "filename" | "content" | "unknown"
    suggested_title: Optional[str] = None

    def to_dict(self) -> dict:
        return {
            "doc_type": self.doc_type,
            "confidence": self.confidence,
            "method": self.method,
            "suggested_title": self.suggested_title,
        }


def classify_by_filename(filename: str) -> Optional[ClassificationResult]:
    """Classify document based on filename patterns."""
    if not filename:
        return None

    for pattern, doc_type in FILENAME_PATTERNS:
        if pattern.search(filename):
            # Generate suggested title from filename
            suggested = _clean_filename_for_title(filename)
            return ClassificationResult(
                doc_type=doc_type,
                confidence=0.95,
                method="filename",
                suggested_title=suggested,
            )
    return None


def classify_by_content(content: str) -> Optional[ClassificationResult]:
    """Classify document based on content patterns."""
    if not content or len(content) < 50:
        return None

    # Score each doc type by matching patterns
    scores: dict[str, float] = {}
    for pattern, doc_type, weight in CONTENT_PATTERNS:
        if pattern.search(content):
            scores[doc_type] = max(scores.get(doc_type, 0), weight)

    if not scores:
        return None

    # Return highest scoring type
    best_type = max(scores, key=lambda k: scores[k])
    return ClassificationResult(
        doc_type=best_type,
        confidence=scores[best_type],
        method="content",
        suggested_title=None,
    )


def classify_document(
    filename: str,
    content_preview: Optional[str] = None,
    bedrock_classification: Optional[ClassificationResult] = None,
) -> ClassificationResult:
    """
    Classify a document by type.

    First attempts filename-based classification (fast path).
    Then checks for Bedrock AI classification (from document parser).
    Falls back to content-based regex classification if both are unavailable.

    Args:
        filename: Original filename of the uploaded document
        content_preview: Optional extracted text content for fallback classification
        bedrock_classification: Optional classification from Bedrock document parser

    Returns:
        ClassificationResult with doc_type, confidence, and method
    """
    # Try filename first (fast path)
    result = classify_by_filename(filename)
    if result:
        logger.info(
            "Classified %s as %s via filename (confidence=%.2f)",
            filename,
            result.doc_type,
            result.confidence,
        )
        return result

    # Use Bedrock AI classification if available and confident
    if bedrock_classification and bedrock_classification.confidence >= 0.7:
        logger.info(
            "Classified %s as %s via bedrock (confidence=%.2f)",
            filename,
            bedrock_classification.doc_type,
            bedrock_classification.confidence,
        )
        return bedrock_classification

    # Fall back to content analysis
    if content_preview:
        result = classify_by_content(content_preview)
        if result:
            result.suggested_title = _clean_filename_for_title(filename)
            logger.info(
                "Classified %s as %s via content (confidence=%.2f)",
                filename,
                result.doc_type,
                result.confidence,
            )
            return result

    # Unknown type
    logger.info("Could not classify %s, defaulting to unknown", filename)
    return ClassificationResult(
        doc_type="unknown",
        confidence=0.0,
        method="unknown",
        suggested_title=_clean_filename_for_title(filename),
    )


def extract_text_preview(
    body: bytes,
    content_type: str,
    max_chars: int = 4000,
) -> Optional[str]:
    """
    Extract text content from uploaded document for classification.

    Args:
        body: Raw file bytes
        content_type: MIME content type
        max_chars: Maximum characters to extract

    Returns:
        Extracted text or None if extraction fails
    """
    try:
        # Plain text / Markdown
        if content_type in ("text/plain", "text/markdown"):
            return body.decode("utf-8", errors="replace")[:max_chars]

        # PDF
        if content_type == "application/pdf":
            return _extract_pdf_text(body, max_chars)

        # DOCX
        if (
            content_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ):
            return _extract_docx_text(body, max_chars)

        # XLSX
        if content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return _extract_xlsx_text(body, max_chars)

        # MS Word (legacy .doc)
        if content_type == "application/msword":
            # Legacy .doc format is harder to parse; skip for now
            return None

        return None

    except Exception as e:
        logger.warning("Text extraction failed for %s: %s", content_type, e)
        return None


def _extract_pdf_text(body: bytes, max_chars: int) -> Optional[str]:
    """Extract text from PDF using pypdf."""
    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning("pypdf not installed, skipping PDF extraction")
        return None

    try:
        reader = PdfReader(io.BytesIO(body))
        text_parts: list[str] = []
        char_count = 0

        for page in reader.pages[:5]:  # Limit to first 5 pages
            page_text = page.extract_text() or ""
            text_parts.append(page_text)
            char_count += len(page_text)
            if char_count >= max_chars:
                break

        full_text = "\n".join(text_parts)
        return full_text[:max_chars]

    except Exception as e:
        logger.warning("PDF text extraction failed: %s", e)
        return None


def _extract_docx_text(body: bytes, max_chars: int) -> Optional[str]:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx not installed, skipping DOCX extraction")
        return None

    try:
        doc = Document(io.BytesIO(body))
        text_parts: list[str] = []
        char_count = 0

        for para in doc.paragraphs:
            text_parts.append(para.text)
            char_count += len(para.text)
            if char_count >= max_chars:
                break

        full_text = "\n".join(text_parts)
        return full_text[:max_chars]

    except Exception as e:
        logger.warning("DOCX text extraction failed: %s", e)
        return None


def _extract_xlsx_text(body: bytes, max_chars: int) -> Optional[str]:
    """Extract text from XLSX using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed, skipping XLSX extraction")
        return None

    try:
        wb = load_workbook(io.BytesIO(body), data_only=True, read_only=True)
        text_parts: list[str] = []
        char_count = 0

        for ws in wb.worksheets:
            if hasattr(ws, "sheet_state") and ws.sheet_state != "visible":
                continue
            text_parts.append(f"Sheet: {ws.title}")
            for row in ws.iter_rows(max_row=50, values_only=True):
                cell_values = [str(c) for c in row if c is not None]
                if cell_values:
                    line = " | ".join(cell_values)
                    text_parts.append(line)
                    char_count += len(line)
                    if char_count >= max_chars:
                        break
            if char_count >= max_chars:
                break

        wb.close()
        full_text = "\n".join(text_parts)
        return full_text[:max_chars]

    except Exception as e:
        logger.warning("XLSX text extraction failed: %s", e)
        return None


def _clean_filename_for_title(filename: str) -> str:
    """Convert a filename to a human-readable title."""
    if not filename:
        return "Uploaded Document"

    filename = filename[:255]

    # Remove extension
    name = filename.rsplit(".", 1)[0] if "." in filename else filename

    # Replace separators with spaces
    name = re.sub(r"[-_]+", " ", name)

    # Remove version numbers like v1, v2, _v3
    parts = [part for part in name.split() if not re.fullmatch(r"(?i)v\d+", part)]
    name = " ".join(parts)

    # Clean up multiple spaces
    name = re.sub(r"\s+", " ", name).strip()

    # Title case
    return name.title() if name else "Uploaded Document"
