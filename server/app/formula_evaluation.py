"""Excel formula calculation helpers for preview rendering.

This module calculates workbook formulas for preview purposes without
modifying the underlying XLSX bytes. Persisted/downloaded workbooks should
retain their original Excel formulas so they continue recalculating in Excel.
"""

from __future__ import annotations

import logging
import os
import tempfile
from typing import Any, Tuple

logger = logging.getLogger("eagle.formula_evaluation")


def _scan_workbook_formulas(xlsx_bytes: bytes) -> tuple[bool, bool]:
    """Return (is_valid_workbook, has_formula_cells)."""
    try:
        from openpyxl import load_workbook
        import io
    except ImportError:
        return False, False

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    except Exception:
        return False, False

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return True, True
    return True, False


def _extract_scalar_value(value: Any) -> Any:
    """Normalize formula-engine outputs to native Python scalars."""
    try:
        import numpy as np
    except ImportError:  # pragma: no cover - numpy is available with formulas
        np = None  # type: ignore[assignment]

    scalar = value.value[0, 0] if hasattr(value, "value") else value

    if isinstance(scalar, tuple):
        if not scalar:
            return None
        if len(scalar) == 1:
            return _extract_scalar_value(scalar[0])
        return scalar[0]

    if np is not None:
        if isinstance(scalar, (np.integer,)):
            return int(scalar)
        if isinstance(scalar, (np.floating,)):
            return float(scalar)
        if isinstance(scalar, np.ndarray):
            return scalar.item()

    if str(scalar) == "empty":
        return None

    return scalar


def _parse_solution_ref(ref: Any) -> tuple[str | None, str | None]:
    """Parse a formulas-engine ref into (sheet_name, cell_ref)."""
    try:
        sheet_part, cell_ref = str(ref).rsplit("!", 1)
        if ":" in cell_ref:
            return None, None
        sheet_name = sheet_part.strip("'")
        if "]" in sheet_name:
            sheet_name = sheet_name.split("]", 1)[1]
        return sheet_name, cell_ref.upper()
    except Exception:
        return None, None


def calculate_workbook_formula_values(
    xlsx_bytes: bytes,
) -> tuple[dict[tuple[str, str], Any], bool]:
    """Calculate workbook formula results without mutating the workbook bytes.

    Returns:
        (formula_values, success)
        formula_values keys are (sheet_name, cell_ref), preserving the actual
        workbook sheet name casing and uppercase cell refs.
    """
    if not xlsx_bytes:
        return {}, False

    is_valid_workbook, has_formula_cells = _scan_workbook_formulas(xlsx_bytes)
    if not is_valid_workbook:
        return {}, False

    if not has_formula_cells:
        return {}, True

    try:
        import formulas
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("formulas/openpyxl not installed, skipping evaluation")
        return {}, False

    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.write(fd, xlsx_bytes)
        os.close(fd)

        logger.info("Loading workbook into formulas engine: %s", tmp_path)
        xl_model = formulas.ExcelModel().loads(tmp_path).finish()
        logger.info("Calculating formulas...")
        solution = xl_model.calculate()
        logger.info("Formula calculation complete, %d results", len(solution))

        wb = load_workbook(tmp_path, data_only=False)
        sheet_lookup = {name.upper(): name for name in wb.sheetnames}
        formula_values: dict[tuple[str, str], Any] = {}

        for ref, value in solution.items():
            sheet_name, cell_ref = _parse_solution_ref(ref)
            if not sheet_name or not cell_ref:
                continue

            actual_name = sheet_lookup.get(sheet_name.upper())
            if not actual_name:
                continue

            try:
                cell = wb[actual_name][cell_ref]
            except Exception:
                continue

            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                continue

            try:
                normalized_value = _extract_scalar_value(value)
            except Exception:
                logger.debug(
                    "Skipping unparseable formula result for %s!%s",
                    actual_name,
                    cell_ref,
                    exc_info=True,
                )
                continue

            formula_values[(actual_name, cell_ref)] = normalized_value

        return formula_values, True

    except Exception as exc:
        logger.warning("Formula evaluation failed: %s", exc, exc_info=False)
        logger.debug("Formula evaluation exception details:", exc_info=True)
        return {}, False
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def evaluate_workbook_formulas(xlsx_bytes: bytes) -> Tuple[bytes, bool]:
    """
    Compatibility wrapper that preserves workbook formulas.

    The returned bytes are the original workbook bytes. Success indicates
    whether formula calculation succeeded for preview purposes.
    """
    _, success = calculate_workbook_formula_values(xlsx_bytes)
    return xlsx_bytes, success


def evaluate_workbook_formulas_safe(xlsx_bytes: bytes) -> bytes:
    """Compatibility wrapper that returns the original workbook bytes."""
    result, _ = evaluate_workbook_formulas(xlsx_bytes)
    return result
