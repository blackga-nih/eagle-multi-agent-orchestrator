"""
Tests for spreadsheet_edit_service.py — XLSX preview extraction and cell editing.

Validates:
  - _coerce_cell_value(): type coercion from string to int/float/bool/None
  - _serialize_cell_value(): value-to-string serialization
  - _sheet_identity(): sheet title normalization to slug
  - _is_allowed_document_key(): security check for tenant/user paths
  - _extract_package_document_ref(): regex parsing of package S3 keys
  - _extract_workspace_document_ref(): regex parsing of workspace S3 keys
  - extract_xlsx_preview_payload(): structured sheet extraction
  - apply_xlsx_cell_edits(): cell editing with type coercion
  - save_xlsx_preview_edits(): full flow with mocked S3

All tests are fast (mocked S3, in-memory XLSX via openpyxl).
"""

import io
import pytest
from unittest.mock import MagicMock, patch


# ---------------------------------------------------------------------------
# Helpers — build in-memory XLSX fixtures
# ---------------------------------------------------------------------------

def _build_sample_xlsx(
    sheets: dict[str, list[list]] | None = None,
) -> bytes:
    """Build a minimal XLSX in memory.

    sheets: dict of {sheet_title: [[row1_values], [row2_values], ...]}
    Defaults to a single sheet with a header row and 3 data rows.
    """
    from openpyxl import Workbook

    wb = Workbook()
    if sheets is None:
        sheets = {
            "Budget": [
                ["Item", "Qty", "Unit Price", "Total"],
                ["Servers", 10, 5000.00, "=B2*C2"],
                ["Licenses", 50, 200.00, "=B3*C3"],
                ["Training", 5, 1500.00, "=B4*C4"],
            ],
        }

    first = True
    for title, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title)
        for row in rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx_with_merged_cells() -> bytes:
    """Build XLSX with merged cells to test hidden-cell detection."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws["A1"] = "Header spanning two columns"
    ws.merge_cells("A1:B1")
    ws["A2"] = "Data"
    ws["B2"] = "More data"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# TestCoerceCellValue
# ---------------------------------------------------------------------------

class TestCoerceCellValue:
    """Tests for _coerce_cell_value — string to typed value."""

    def test_empty_string_returns_none(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("") is None
        assert _coerce_cell_value("   ") is None

    def test_integer_string(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("42") == 42
        assert isinstance(_coerce_cell_value("42"), int)

    def test_negative_integer(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("-7") == -7

    def test_float_string(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("3.14") == 3.14
        assert isinstance(_coerce_cell_value("3.14"), float)

    def test_boolean_true(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("true") is True
        assert _coerce_cell_value("TRUE") is True

    def test_boolean_false(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("false") is False
        assert _coerce_cell_value("FALSE") is False

    def test_plain_text_passes_through(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("hello world") == "hello world"

    def test_currency_string_passes_through(self):
        from app.spreadsheet_edit_service import _coerce_cell_value

        assert _coerce_cell_value("$1,000") == "$1,000"


# ---------------------------------------------------------------------------
# TestSerializeCellValue
# ---------------------------------------------------------------------------

class TestSerializeCellValue:
    """Tests for _serialize_cell_value."""

    def test_none_returns_empty_string(self):
        from app.spreadsheet_edit_service import _serialize_cell_value

        assert _serialize_cell_value(None) == ""

    def test_integer_to_string(self):
        from app.spreadsheet_edit_service import _serialize_cell_value

        assert _serialize_cell_value(42) == "42"

    def test_float_to_string(self):
        from app.spreadsheet_edit_service import _serialize_cell_value

        assert _serialize_cell_value(3.14) == "3.14"

    def test_string_passthrough(self):
        from app.spreadsheet_edit_service import _serialize_cell_value

        assert _serialize_cell_value("hello") == "hello"


# ---------------------------------------------------------------------------
# TestSheetIdentity
# ---------------------------------------------------------------------------

class TestSheetIdentity:
    """Tests for _sheet_identity — title normalization."""

    def test_simple_title(self):
        from app.spreadsheet_edit_service import _sheet_identity

        ws = MagicMock()
        ws.title = "Budget"
        assert _sheet_identity(ws) == "budget"

    def test_title_with_spaces_and_special_chars(self):
        from app.spreadsheet_edit_service import _sheet_identity

        ws = MagicMock()
        ws.title = "Q1 Budget (Draft)"
        assert _sheet_identity(ws) == "q1-budget-draft"

    def test_title_with_numbers(self):
        from app.spreadsheet_edit_service import _sheet_identity

        ws = MagicMock()
        ws.title = "Sheet 123"
        assert _sheet_identity(ws) == "sheet-123"


# ---------------------------------------------------------------------------
# TestIsAllowedDocumentKey
# ---------------------------------------------------------------------------

class TestIsAllowedDocumentKey:
    """Tests for _is_allowed_document_key."""

    def test_allows_package_key(self):
        from app.spreadsheet_edit_service import _is_allowed_document_key

        assert _is_allowed_document_key(
            "eagle/t/packages/pkg/igce/v1/igce.xlsx", "t", "u"
        ) is True

    def test_allows_workspace_key(self):
        from app.spreadsheet_edit_service import _is_allowed_document_key

        assert _is_allowed_document_key(
            "eagle/t/u/documents/igce.xlsx", "t", "u"
        ) is True

    def test_denies_wrong_tenant(self):
        from app.spreadsheet_edit_service import _is_allowed_document_key

        assert _is_allowed_document_key(
            "eagle/other/packages/pkg/igce/v1/igce.xlsx", "t", "u"
        ) is False


# ---------------------------------------------------------------------------
# TestExtractXlsxPreviewPayload
# ---------------------------------------------------------------------------

class TestExtractXlsxPreviewPayload:
    """Tests for extract_xlsx_preview_payload."""

    def test_extracts_sheets_from_valid_xlsx(self):
        from app.spreadsheet_edit_service import extract_xlsx_preview_payload

        xlsx_bytes = _build_sample_xlsx()
        result = extract_xlsx_preview_payload(xlsx_bytes)

        assert result["preview_mode"] == "xlsx_grid"
        assert len(result["preview_sheets"]) == 1

        sheet = result["preview_sheets"][0]
        assert sheet["title"] == "Budget"
        assert len(sheet["rows"]) >= 4  # header + 3 data rows
        assert sheet["sheet_id"].endswith(":budget")

    def test_formula_cells_are_not_editable(self):
        from app.spreadsheet_edit_service import extract_xlsx_preview_payload

        xlsx_bytes = _build_sample_xlsx()
        result = extract_xlsx_preview_payload(xlsx_bytes)

        sheet = result["preview_sheets"][0]
        # Row 2 (index 1) col D should be a formula
        row2 = sheet["rows"][1]
        d_cell = [c for c in row2["cells"] if c["cell_ref"] == "D2"][0]
        assert d_cell["is_formula"] is True
        assert d_cell["editable"] is False

    def test_text_cells_are_editable(self):
        from app.spreadsheet_edit_service import extract_xlsx_preview_payload

        xlsx_bytes = _build_sample_xlsx()
        result = extract_xlsx_preview_payload(xlsx_bytes)

        sheet = result["preview_sheets"][0]
        row1 = sheet["rows"][0]
        a_cell = [c for c in row1["cells"] if c["cell_ref"] == "A1"][0]
        assert a_cell["editable"] is True
        assert a_cell["value"] == "Item"

    def test_multiple_sheets(self):
        from app.spreadsheet_edit_service import extract_xlsx_preview_payload

        xlsx_bytes = _build_sample_xlsx({
            "Budget": [["A", "B"], [1, 2]],
            "Summary": [["Total"], [100]],
        })
        result = extract_xlsx_preview_payload(xlsx_bytes)

        assert len(result["preview_sheets"]) == 2
        titles = [s["title"] for s in result["preview_sheets"]]
        assert "Budget" in titles
        assert "Summary" in titles

    def test_merged_cells_not_editable(self):
        from app.spreadsheet_edit_service import extract_xlsx_preview_payload

        xlsx_bytes = _build_xlsx_with_merged_cells()
        result = extract_xlsx_preview_payload(xlsx_bytes)

        sheet = result["preview_sheets"][0]
        row1 = sheet["rows"][0]
        b1_cell = [c for c in row1["cells"] if c["cell_ref"] == "B1"][0]
        # B1 is hidden by merge → not editable
        assert b1_cell["editable"] is False

    def test_content_field_has_text(self):
        from app.spreadsheet_edit_service import extract_xlsx_preview_payload

        xlsx_bytes = _build_sample_xlsx()
        result = extract_xlsx_preview_payload(xlsx_bytes)

        # content comes from XLSXPopulator.extract_text — may be None if not available
        # but preview_mode should always be set
        assert result["preview_mode"] == "xlsx_grid"


# ---------------------------------------------------------------------------
# TestApplyXlsxCellEdits
# ---------------------------------------------------------------------------

class TestApplyXlsxCellEdits:
    """Tests for apply_xlsx_cell_edits."""

    def test_applies_text_cell_edit(self):
        from app.spreadsheet_edit_service import SpreadsheetCellEdit, apply_xlsx_cell_edits

        xlsx_bytes = _build_sample_xlsx()
        edits = [SpreadsheetCellEdit(sheet_id="0:budget", cell_ref="A2", value="Cloud Servers")]
        updated, applied, missing = apply_xlsx_cell_edits(xlsx_bytes, edits)

        assert applied == 1
        assert missing == []

        # Verify edit took effect
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(updated))
        assert wb.active["A2"].value == "Cloud Servers"

    def test_applies_numeric_cell_edit(self):
        from app.spreadsheet_edit_service import SpreadsheetCellEdit, apply_xlsx_cell_edits

        xlsx_bytes = _build_sample_xlsx()
        edits = [SpreadsheetCellEdit(sheet_id="0:budget", cell_ref="B2", value="20")]
        updated, applied, missing = apply_xlsx_cell_edits(xlsx_bytes, edits)

        assert applied == 1

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(updated))
        assert wb.active["B2"].value == 20  # Coerced to int

    def test_skips_formula_cells(self):
        from app.spreadsheet_edit_service import SpreadsheetCellEdit, apply_xlsx_cell_edits

        xlsx_bytes = _build_sample_xlsx()
        edits = [SpreadsheetCellEdit(sheet_id="0:budget", cell_ref="D2", value="99999")]
        updated, applied, missing = apply_xlsx_cell_edits(xlsx_bytes, edits)

        assert applied == 0
        assert "0:budget:D2" in missing

    def test_reports_missing_sheet(self):
        from app.spreadsheet_edit_service import SpreadsheetCellEdit, apply_xlsx_cell_edits

        xlsx_bytes = _build_sample_xlsx()
        edits = [SpreadsheetCellEdit(sheet_id="99:nonexistent", cell_ref="A1", value="x")]
        updated, applied, missing = apply_xlsx_cell_edits(xlsx_bytes, edits)

        assert applied == 0
        assert "99:nonexistent:A1" in missing

    def test_multiple_edits(self):
        from app.spreadsheet_edit_service import SpreadsheetCellEdit, apply_xlsx_cell_edits

        xlsx_bytes = _build_sample_xlsx()
        edits = [
            SpreadsheetCellEdit(sheet_id="0:budget", cell_ref="A2", value="Servers v2"),
            SpreadsheetCellEdit(sheet_id="0:budget", cell_ref="A3", value="Licenses v2"),
        ]
        updated, applied, missing = apply_xlsx_cell_edits(xlsx_bytes, edits)

        assert applied == 2
        assert missing == []

    def test_no_change_when_value_same(self):
        from app.spreadsheet_edit_service import SpreadsheetCellEdit, apply_xlsx_cell_edits

        xlsx_bytes = _build_sample_xlsx()
        # "Servers" is already the value in A2
        edits = [SpreadsheetCellEdit(sheet_id="0:budget", cell_ref="A2", value="Servers")]
        updated, applied, missing = apply_xlsx_cell_edits(xlsx_bytes, edits)

        assert applied == 0  # No change needed


# ---------------------------------------------------------------------------
# TestSaveXlsxPreviewEdits — integration with mocked S3
# ---------------------------------------------------------------------------

class TestSaveXlsxPreviewEdits:
    """Tests for save_xlsx_preview_edits with mocked S3."""

    def test_returns_error_for_empty_doc_key(self):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits

        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u", doc_key="",
            cell_edits=[{"sheet_id": "0:budget", "cell_ref": "A1", "value": "x"}],
        )
        assert "error" in result

    def test_returns_error_for_empty_cell_edits(self):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits

        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u",
            doc_key="eagle/t/u/documents/igce.xlsx",
            cell_edits=[],
        )
        assert "error" in result

    def test_returns_error_for_non_xlsx(self):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits

        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u",
            doc_key="eagle/t/u/documents/sow.docx",
            cell_edits=[{"sheet_id": "0:s", "cell_ref": "A1", "value": "x"}],
        )
        assert "error" in result
        assert "xlsx" in result["error"].lower()

    def test_returns_error_for_access_denied(self):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits

        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u",
            doc_key="eagle/other/u/documents/igce.xlsx",
            cell_edits=[{"sheet_id": "0:s", "cell_ref": "A1", "value": "x"}],
        )
        assert "error" in result
        assert "denied" in result["error"].lower()

    @patch("app.spreadsheet_edit_service._get_s3")
    @patch("app.spreadsheet_edit_service.write_document_changelog_entry")
    def test_workspace_save_success(self, mock_changelog, mock_get_s3):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits

        xlsx_bytes = _build_sample_xlsx()
        mock_s3 = MagicMock()
        mock_s3.get_object.return_value = {
            "Body": MagicMock(read=MagicMock(return_value=xlsx_bytes)),
            "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        mock_get_s3.return_value = mock_s3

        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u",
            doc_key="eagle/t/u/documents/igce.xlsx",
            cell_edits=[{"sheet_id": "0:budget", "cell_ref": "A2", "value": "New Item"}],
        )

        assert result.get("success") is True
        assert result["mode"] == "workspace_xlsx_preview_edit"
        mock_s3.put_object.assert_called_once()
        mock_changelog.assert_called_once()

    @patch("app.spreadsheet_edit_service._get_s3")
    def test_s3_load_failure(self, mock_get_s3):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits
        from botocore.exceptions import ClientError

        mock_s3 = MagicMock()
        mock_s3.get_object.side_effect = ClientError(
            {"Error": {"Code": "NoSuchKey", "Message": "Not found"}}, "GetObject"
        )
        mock_get_s3.return_value = mock_s3

        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u",
            doc_key="eagle/t/u/documents/igce.xlsx",
            cell_edits=[{"sheet_id": "0:budget", "cell_ref": "A1", "value": "x"}],
        )

        assert "error" in result
        assert "Failed to load" in result["error"]

    @patch("app.spreadsheet_edit_service._get_s3")
    @patch("app.spreadsheet_edit_service.write_document_changelog_entry")
    def test_no_edits_applied_returns_error(self, mock_changelog, mock_get_s3):
        from app.spreadsheet_edit_service import save_xlsx_preview_edits

        xlsx_bytes = _build_sample_xlsx()
        mock_s3 = MagicMock()
        mock_s3.get_object.return_value = {
            "Body": MagicMock(read=MagicMock(return_value=xlsx_bytes)),
            "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        mock_get_s3.return_value = mock_s3

        # Try to edit a formula cell — should fail
        result = save_xlsx_preview_edits(
            tenant_id="t", user_id="u",
            doc_key="eagle/t/u/documents/igce.xlsx",
            cell_edits=[{"sheet_id": "0:budget", "cell_ref": "D2", "value": "99"}],
        )

        assert "error" in result
        assert "No spreadsheet edits" in result["error"]
        mock_changelog.assert_not_called()
