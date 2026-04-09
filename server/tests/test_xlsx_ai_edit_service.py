"""Tests for AI-assisted IGCE XLSX edit service."""

from __future__ import annotations

import io
import os
import sys
from unittest.mock import MagicMock, patch

_server_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
if _server_dir not in sys.path:
    sys.path.insert(0, _server_dir)

from app.spreadsheet_edit_service import extract_xlsx_preview_payload  # noqa: E402
from app.xlsx_ai_edit_service import edit_igce_xlsx_document  # noqa: E402


def _build_commercial_igce_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "IGCE"
    ws["A7"] = "Cloud Architect"
    ws["C5"] = "12 months"
    ws["C7"] = 2080
    ws["E7"] = 175
    ws["G7"] = "=C7*E7"
    ws["A30"] = "AWS Licensing"
    ws["E30"] = 180000

    ws_services = wb.create_sheet("IT Services")
    ws_services["B5"] = "Firm-Fixed-Price"
    ws_services["B6"] = "2026-01-01"
    ws_services["D6"] = "2026-12-31"
    ws_services["A12"] = "Cloud Architect"
    ws_services["B12"] = 2080
    ws_services["C12"] = 175
    ws_services["D12"] = "=B12*C12"
    ws_services["A13"] = "Manager II"
    ws_services["B13"] = 1872
    ws_services["C13"] = ""
    ws_services["D13"] = "=B13*C13"

    ws_goods = wb.create_sheet("IT Goods")
    ws_goods["B5"] = "Firm-Fixed-Price"
    ws_goods["B6"] = "2026-09-30"
    ws_goods["A10"] = "AWS Licensing"
    ws_goods["E10"] = 12
    ws_goods["F10"] = 15000
    ws_goods["G10"] = "=F10*E10"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_partial_context_fill_xlsx() -> bytes:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(_build_commercial_igce_xlsx()))

    ws = wb["IGCE"]
    ws["A8"] = "DevOps Engineer"
    ws["C8"] = None
    ws["E8"] = None
    ws["C5"] = None

    ws_services = wb["IT Services"]
    ws_services["A13"] = "DevOps Engineer"
    ws_services["B13"] = None
    ws_services["C13"] = None
    ws_services["B5"] = None

    ws_goods = wb["IT Goods"]
    ws_goods["B5"] = None

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_ige_products_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "INDEPENDENT GOVERNMENT ESTIMATE (IGE) FOR PRODUCTS"
    ws["B3"] = "Products (including software, licenses)"
    ws["A4"] = "Item #"
    ws["B4"] = "Manufacturer/Description/Part/Model Number*"
    ws["C4"] = "Qty"
    ws["D4"] = "Price"
    ws["E4"] = "Extended Amount"
    ws["A5"] = 1
    ws["B5"] = "Microscope / Acme / M-100"
    ws["C5"] = 2
    ws["D5"] = 5000
    ws["E5"] = "=SUM(C5*D5)"
    for row in (6, 7, 8, 9):
        ws[f"E{row}"] = f"=SUM(C{row}*D{row})"
    ws["B10"] = "Shipping (include as needed)"
    ws["C10"] = 1
    ws["D10"] = 400
    ws["E10"] = "=SUM(C10*D10)"
    ws["B11"] = "Installation (include as needed)"
    ws["E11"] = "=SUM(C11*D11)"
    ws["B12"] = "Training (include as needed)"
    ws["E13"] = "=SUM(C13*D13)"
    ws["B14"] = "  SUBTOTAL (Products)"
    ws["E14"] = "=SUM(E5:E13)"
    ws["B16"] = "TOTAL IGE"
    ws["E16"] = "=E14"
    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_ige_services_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    base = wb.active
    base.title = "Base Period"
    base["A1"] = "INDEPENDENT GOVERNMENT ESTIMATE (IGE) FOR SERVICES"
    base["A4"] = "Brief Description of Service*"
    base["C4"] = "Period of Performance\n(Base)"
    base.merge_cells("C4:D4")
    base["A5"] = "Cloud Hosting"
    base["C5"] = 12
    base["D5"] = 5000
    base["E5"] = "=SUM(C5*D5)"
    for row in (6, 7, 8, 9, 10, 11, 12, 13):
        base[f"E{row}"] = f"=SUM(C{row}*D{row})"
    base["E14"] = "=SUM(E5:E13)"
    base["E16"] = "=E14"

    option = wb.create_sheet("Option Period One")
    option["A1"] = "INDEPENDENT GOVERNMENT COST ESTIMATE (IGCE)"
    option["A4"] = "Brief Description of Service*"
    option["C4"] = "Period of Performance\n(Option Period 1)"
    option.merge_cells("C4:D4")
    option["A5"] = "Cloud Hosting Option 1"
    option["C5"] = 12
    option["D5"] = 5500
    option["E5"] = "=SUM(C5*D5)"
    for row in (6, 7, 8, 9, 10, 11, 12, 13):
        option[f"E{row}"] = f"=SUM(C{row}*D{row})"
    option["E14"] = "=SUM(E5:E13)"
    option["E16"] = "=E14"

    wb.create_sheet("Sheet3")
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_edit_igce_xlsx_document_updates_labor_rate():
    xlsx_bytes = _build_commercial_igce_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_document",
        return_value={"title": "Cloud Migration IGCE", "session_id": "sess-123"},
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Cloud hosting support"},
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 2,
            "s3_key": "eagle/dev-tenant/packages/PKG-1/igce/v2/IGCE.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/packages/PKG-1/igce/v1/IGCE.xlsx",
            request_text="Set Cloud Architect to $190/hour",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert {"sheet_id": "0:igce", "cell_ref": "E7", "value": "190"} in sent_edits
    assert {"sheet_id": "1:it-services", "cell_ref": "C12", "value": "190"} in sent_edits
    assert "Cloud Architect rate" in result["assistant_message"]


def test_edit_igce_xlsx_document_supports_workbook_wide_it_services_row_edit():
    xlsx_bytes = _build_commercial_igce_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_document",
        return_value={"title": "Cloud Migration IGCE", "session_id": "sess-123"},
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Cloud hosting support"},
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 2,
            "s3_key": "eagle/dev-tenant/packages/PKG-1/igce/v2/IGCE.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/packages/PKG-1/igce/v1/IGCE.xlsx",
            request_text="Set the IT Services Manager II hourly rate to 175/hr",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert {"sheet_id": "1:it-services", "cell_ref": "C13", "value": "175"} in sent_edits
    assert "Manager II rate" in result["assistant_message"]


def test_edit_igce_xlsx_document_supports_direct_cell_edit_request():
    xlsx_bytes = _build_commercial_igce_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_document",
        return_value={"title": "Cloud Migration IGCE", "session_id": "sess-123"},
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Cloud hosting support"},
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 2,
            "s3_key": "eagle/dev-tenant/packages/PKG-1/igce/v2/IGCE.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/packages/PKG-1/igce/v1/IGCE.xlsx",
            request_text="Set IT Goods F10 to 12000",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert {"sheet_id": "2:it-goods", "cell_ref": "F10", "value": "12000"} in sent_edits
    assert "IT Goods F10" in result["assistant_message"]


def test_edit_igce_xlsx_document_supports_ige_products_updates():
    xlsx_bytes = _build_ige_products_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.find_document_by_s3_key",
        create=True,
        return_value=None,
    ), patch(
        "app.user_document_store.find_document_by_s3_key",
        return_value={
            "document_id": "doc-prod-1",
            "title": "IGE Products",
            "doc_type": "igce",
            "session_id": "sess-123",
            "template_id": "eagle-knowledge-base/approved/supervisor-core/essential-templates/4.a. IGE for Products.xlsx",
        },
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Lab equipment"},
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 0,
            "s3_key": "eagle/dev-tenant/dev-user/documents/ige_products.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/dev-user/documents/ige_products.xlsx",
            request_text="Set microscope quantity to 3",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert {"sheet_id": "0:sheet1", "cell_ref": "C5", "value": "3"} in sent_edits
    assert "Microscope / Acme / M-100 quantity" in result["assistant_message"]


def test_edit_igce_xlsx_document_supports_ige_services_updates():
    xlsx_bytes = _build_ige_services_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.user_document_store.find_document_by_s3_key",
        return_value={
            "document_id": "doc-svc-1",
            "title": "IGE Services",
            "doc_type": "igce",
            "session_id": "sess-456",
            "template_id": "eagle-knowledge-base/approved/supervisor-core/essential-templates/4.b. IGE for Services based on Catalog Price.xlsx",
        },
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Cloud hosting services"},
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 0,
            "s3_key": "eagle/dev-tenant/dev-user/documents/ige_services.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/dev-user/documents/ige_services.xlsx",
            request_text="Set cloud hosting unit price to 6000",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert {"sheet_id": "0:base-period", "cell_ref": "D5", "value": "6000"} in sent_edits
    assert "Base Period Cloud Hosting unit price" in result["assistant_message"]


def test_edit_igce_xlsx_document_uses_ai_extractor_for_contextual_request():
    xlsx_bytes = _build_commercial_igce_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}
    bedrock = MagicMock()
    bedrock.converse.return_value = {
        "output": {
            "message": {
                "content": [
                    {
                        "text": (
                            '{"action":"apply","clarification":"",'
                            '"intents":[{"type":"update_contract_type","value":"Time & Materials"}]}'
                        )
                    }
                ]
            }
        }
    }

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_document",
        return_value={"title": "Cloud Migration IGCE", "session_id": "sess-123"},
    ), patch(
        "app.xlsx_ai_edit_service.get_package",
        return_value={"package_id": "PKG-1", "title": "Cloud package"},
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Use T&M as discussed in chat"},
    ), patch(
        "app.xlsx_ai_edit_service._get_doc_gen_bedrock",
        return_value=bedrock,
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 2,
            "s3_key": "eagle/dev-tenant/packages/PKG-1/igce/v2/IGCE.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/packages/PKG-1/igce/v1/IGCE.xlsx",
            request_text="Make it T&M like we discussed earlier.",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert {"sheet_id": "1:it-services", "cell_ref": "B5", "value": "Time & Materials"} in sent_edits
    assert {"sheet_id": "2:it-goods", "cell_ref": "B5", "value": "Time & Materials"} in sent_edits
    assert result["origin_context_available"] is True


def test_edit_igce_xlsx_context_fill_with_source_data():
    """Test context-fill request uses stored source_data to fill empty cells."""
    xlsx_bytes = _build_partial_context_fill_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    # Document has source_data from original generation
    source_data = {
        "line_items": [
            {"description": "Cloud Architect", "rate": 200, "hours": 1500},
            {"description": "DevOps Engineer", "rate": 175, "hours": 1000},
        ],
        "goods_items": [
            {"product_name": "AWS Licensing", "quantity": 24, "unit_price": 12000},
        ],
        "contract_type": "T&M",
        "period_months": 18,
    }

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_document",
        return_value={
            "title": "Cloud Migration IGCE",
            "session_id": "sess-123",
            "source_data": source_data,
        },
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={},
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 2,
            "s3_key": "eagle/dev-tenant/packages/PKG-1/igce/v2/IGCE.xlsx",
        },
    ):
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/packages/PKG-1/igce/v1/IGCE.xlsx",
            request_text="Fill in the rest from our earlier conversation",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    # Should have skipped_fields for values that couldn't be filled
    assert "skipped_fields" in result
    # Should report what was updated in the message
    assert "Updated" in result["assistant_message"]


def test_edit_igce_xlsx_workspace_document_uses_unified_metadata():
    xlsx_bytes = _build_partial_context_fill_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={"description": "Cloud hosting support"},
    ), patch(
        "app.user_document_store.find_document_by_s3_key",
        return_value={
            "document_id": "doc-123",
            "title": "Workspace IGCE",
            "doc_type": "igce",
            "session_id": "sess-123",
            "source_data": {
                "line_items": [{"description": "DevOps Engineer", "rate": 190, "hours": 2200}],
                "contract_type": "FFP",
            },
        },
    ), patch(
        "app.xlsx_ai_edit_service.save_xlsx_preview_edits",
        return_value={
            "success": True,
            "preview_mode": "xlsx_grid",
            "preview_sheets": preview["preview_sheets"],
            "content": preview["content"],
            "version": 0,
            "s3_key": "eagle/dev-tenant/dev-user/documents/igce_20260402.xlsx",
        },
    ) as save_mock:
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/dev-user/documents/igce_20260402.xlsx",
            request_text="Fill in the rest from our earlier conversation",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is False
    assert result["origin_context_available"] is True
    sent_edits = save_mock.call_args.kwargs["cell_edits"]
    assert any(edit["cell_ref"] == "E8" and edit["value"] == "190" for edit in sent_edits)
    assert any(edit["cell_ref"] == "B5" and edit["value"] == "FFP" for edit in sent_edits)


def test_edit_igce_xlsx_context_fill_no_source_data():
    """Test context-fill request fails gracefully when no source_data is available."""
    xlsx_bytes = _build_commercial_igce_xlsx()
    s3 = MagicMock()
    s3.get_object.return_value = {"Body": io.BytesIO(xlsx_bytes)}

    with patch("app.xlsx_ai_edit_service.get_s3", return_value=s3), patch(
        "app.xlsx_ai_edit_service.get_document",
        return_value={
            "title": "Cloud Migration IGCE",
            "session_id": None,
            # No source_data
        },
    ), patch(
        "app.xlsx_ai_edit_service.get_package", return_value=None
    ), patch(
        "app.xlsx_ai_edit_service._augment_document_data_from_context",
        return_value={},
    ):
        result = edit_igce_xlsx_document(
            tenant_id="dev-tenant",
            user_id="dev-user",
            doc_key="eagle/dev-tenant/packages/PKG-1/igce/v1/IGCE.xlsx",
            request_text="Complete the IGCE using our discussion",
        )

    assert result["success"] is True
    assert result["clarification_needed"] is True
    assert "context" in result["assistant_message"].lower()
    assert result["origin_context_available"] is False


def test_context_fill_request_detection():
    """Test that context-fill patterns are correctly detected."""
    from app.igce_xlsx_edit_resolver import resolve_igce_edit_request, build_commercial_igce_workbook_context

    xlsx_bytes = _build_commercial_igce_xlsx()
    preview = extract_xlsx_preview_payload(xlsx_bytes)
    workbook = build_commercial_igce_workbook_context(preview["preview_sheets"])

    # These should all be detected as context-fill requests
    context_fill_phrases = [
        "Fill from context",
        "Complete the IGCE",
        "Populate using our conversation",
        "Use our earlier discussion to fill in the blanks",
        "Fill in the rest",
        "Auto fill the workbook",
    ]

    for phrase in context_fill_phrases:
        result = resolve_igce_edit_request(phrase, workbook)
        assert result.is_context_fill_request, f"Failed for: {phrase}"
        assert not result.intents
        assert not result.clarification
