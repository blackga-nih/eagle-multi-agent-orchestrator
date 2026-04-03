"""Tests for Excel formula evaluation."""

import io
import os
import sys

# Add server directory to path for imports
_server_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
if _server_dir not in sys.path:
    sys.path.insert(0, _server_dir)

from openpyxl import Workbook  # noqa: E402

from app.formula_evaluation import calculate_workbook_formula_values, evaluate_workbook_formulas, evaluate_workbook_formulas_safe  # noqa: E402


def _build_formula_xlsx() -> bytes:
    """Create test workbook with various formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"

    # Simple values
    ws["A1"] = 10
    ws["B1"] = 20
    ws["C1"] = "=A1+B1"  # Should evaluate to 30

    ws["A2"] = 5
    ws["B2"] = "=A2*2"  # Should evaluate to 10
    ws["C2"] = "=SUM(A1:B2)"  # Should evaluate to 45 (10+20+5+10)

    # More complex formulas
    ws["A3"] = 100
    ws["B3"] = 0.15
    ws["C3"] = "=A3*B3"  # Should evaluate to 15

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_simple_xlsx() -> bytes:
    """Create test workbook without formulas."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Item 1"
    ws["B2"] = 100

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class TestEvaluateWorkbookFormulas:
    """Tests for evaluate_workbook_formulas function."""

    def test_reports_success_for_formula_workbook(self):
        """Formula workbooks should report successful preview calculation."""
        xlsx_bytes = _build_formula_xlsx()
        result_bytes, success = evaluate_workbook_formulas(xlsx_bytes)

        assert success is True
        assert result_bytes == xlsx_bytes

    def test_preserves_formulas_in_workbook(self):
        """Formula cells should remain formulas after evaluation."""
        xlsx_bytes = _build_formula_xlsx()
        result_bytes, _ = evaluate_workbook_formulas(xlsx_bytes)

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(result_bytes), data_only=False)
        ws = wb.active

        assert ws["C1"].value == "=A1+B1"
        assert ws["B2"].value == "=A2*2"
        assert ws["C2"].value == "=SUM(A1:B2)"

    def test_handles_workbook_without_formulas(self):
        """Workbooks without formulas should pass through unchanged."""
        xlsx_bytes = _build_simple_xlsx()
        result_bytes, success = evaluate_workbook_formulas(xlsx_bytes)

        assert success is True

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(result_bytes), data_only=True)
        ws = wb.active

        assert ws["A1"].value == "Name"
        assert ws["B2"].value == 100


class TestCalculateWorkbookFormulaValues:
    """Tests for preview-only formula calculations."""

    def test_calculates_simple_addition(self):
        """Formula =A1+B1 should evaluate to the sum of cells."""
        formula_values, success = calculate_workbook_formula_values(_build_formula_xlsx())

        assert success is True
        assert formula_values[("Calculations", "C1")] == 30

    def test_calculates_multiplication(self):
        """Formula =A2*2 should evaluate correctly."""
        formula_values, success = calculate_workbook_formula_values(_build_formula_xlsx())

        assert success is True
        assert formula_values[("Calculations", "B2")] == 10

    def test_calculates_sum_function(self):
        """Formula =SUM(A1:B2) should evaluate correctly."""
        formula_values, success = calculate_workbook_formula_values(_build_formula_xlsx())

        assert success is True
        assert formula_values[("Calculations", "C2")] == 45

    def test_handles_invalid_xlsx(self):
        """Invalid XLSX should return original bytes with success=False."""
        invalid_bytes = b"not a valid xlsx file"
        result, success = calculate_workbook_formula_values(invalid_bytes)

        assert success is False
        assert result == {}

    def test_handles_empty_bytes(self):
        """Empty bytes should return original with success=False."""
        empty_bytes = b""
        result, success = calculate_workbook_formula_values(empty_bytes)

        assert success is False
        assert result == {}


class TestEvaluateWorkbookFormulasSafe:
    """Tests for the safe wrapper function."""

    def test_returns_original_bytes_on_success(self):
        """Should return original workbook bytes when successful."""
        xlsx_bytes = _build_formula_xlsx()
        result_bytes = evaluate_workbook_formulas_safe(xlsx_bytes)
        assert result_bytes == xlsx_bytes

    def test_returns_original_bytes_on_failure(self):
        """Should return original bytes on failure without raising."""
        invalid_bytes = b"invalid xlsx"
        result_bytes = evaluate_workbook_formulas_safe(invalid_bytes)

        # Should return original, not raise
        assert result_bytes == invalid_bytes


class TestIGCEStyleFormulas:
    """Tests simulating IGCE spreadsheet patterns."""

    def test_line_item_totals(self):
        """IGCE-style qty * price = total formulas should work."""
        wb = Workbook()
        ws = wb.active
        ws.title = "IGCE"

        # Header row
        ws["A1"] = "Item"
        ws["B1"] = "Qty"
        ws["C1"] = "Unit Price"
        ws["D1"] = "Total"

        # Line items with formulas
        ws["A2"] = "Microscope"
        ws["B2"] = 2
        ws["C2"] = 1500
        ws["D2"] = "=B2*C2"  # Should be 3000

        ws["A3"] = "Centrifuge"
        ws["B3"] = 1
        ws["C3"] = 5000
        ws["D3"] = "=B3*C3"  # Should be 5000

        # Grand total
        ws["D4"] = "=SUM(D2:D3)"  # Should be 8000

        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()

        # Evaluate
        formula_values, success = calculate_workbook_formula_values(xlsx_bytes)
        assert success is True
        assert formula_values[("IGCE", "D2")] == 3000
        assert formula_values[("IGCE", "D3")] == 5000
        assert formula_values[("IGCE", "D4")] == 8000
