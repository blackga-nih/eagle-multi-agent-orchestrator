"""Tests for 2026-04-08 dev huddle fixes.

Covers:
- Plan 1: Streaming noise reduction (tool_input_delta filtering, tool_input dedup)
- Plan 3: Statement of Need (SON) template generation pipeline
- Plan 4: Package checklist endpoint (backend portion)
- Plan 5: XLSX document page renders markdown above spreadsheet
- Plan 8: Baseline questions skill updated for 14-question suite

Run: pytest server/tests/test_huddle_20260408_fixes.py -v
"""

import asyncio
import json
import os
import sys
from functools import wraps
from unittest.mock import patch

import pytest

_server_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
if _server_dir not in sys.path:
    sys.path.insert(0, _server_dir)


# ── Helpers ──────────────────────────────────────────────────────────


def _parse_sse_events(raw_lines: list[str]) -> list[dict]:
    """Parse SSE strings into dicts, skipping keepalives and blanks."""
    events = []
    for line in raw_lines:
        line = line.strip()
        if line.startswith("data: "):
            try:
                events.append(json.loads(line[6:]))
            except json.JSONDecodeError:
                pass
    return events


async def _collect_stream(gen) -> list[str]:
    lines = []
    async for line in gen:
        lines.append(line)
    return lines


def async_test(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        return asyncio.run(fn(*args, **kwargs))
    return wrapper


# ══════════════════════════════════════════════════════════════════════
# PLAN 1 — Streaming noise reduction
# ══════════════════════════════════════════════════════════════════════


class TestToolInputDeltaFiltering:
    """tool_input_delta should only pass through for document-creation tools."""

    @async_test
    async def test_create_document_delta_passes_through(self):
        """tool_input_delta for create_document should emit an SSE event."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            yield {"type": "tool_use", "name": "create_document", "input": {}, "tool_use_id": "tu-1"}
            yield {"type": "tool_input_delta", "tool_use_id": "tu-1", "delta": '{"content": "# SOW"}', "name": "create_document"}
            yield {"type": "complete", "text": "", "tools_called": ["create_document"], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="create a sow",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        deltas = [e for e in events if e.get("type") == "tool_input_delta"]
        assert len(deltas) == 1
        assert deltas[0]["metadata"]["tool_name"] == "create_document"

    @async_test
    async def test_update_document_delta_passes_through(self):
        """tool_input_delta for update_document should emit an SSE event."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            yield {"type": "tool_use", "name": "update_document", "input": {}, "tool_use_id": "tu-2"}
            yield {"type": "tool_input_delta", "tool_use_id": "tu-2", "delta": "partial", "name": "update_document"}
            yield {"type": "complete", "text": "", "tools_called": ["update_document"], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="update doc",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        deltas = [e for e in events if e.get("type") == "tool_input_delta"]
        assert len(deltas) == 1

    @async_test
    async def test_knowledge_search_delta_filtered(self):
        """tool_input_delta for knowledge_search should NOT emit (noise reduction)."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            yield {"type": "tool_use", "name": "knowledge_search", "input": {}, "tool_use_id": "tu-3"}
            yield {"type": "tool_input_delta", "tool_use_id": "tu-3", "delta": '{"query": "FAR"}', "name": "knowledge_search"}
            yield {"type": "tool_result", "name": "knowledge_search", "result": {"text": "found"}}
            yield {"type": "complete", "text": "done", "tools_called": ["knowledge_search"], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="search",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        deltas = [e for e in events if e.get("type") == "tool_input_delta"]
        assert len(deltas) == 0, "knowledge_search deltas should be filtered out"

    @async_test
    async def test_compliance_matrix_delta_filtered(self):
        """tool_input_delta for query_compliance_matrix should NOT emit."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            yield {"type": "tool_use", "name": "query_compliance_matrix", "input": {}, "tool_use_id": "tu-4"}
            yield {"type": "tool_input_delta", "tool_use_id": "tu-4", "delta": "partial", "name": "query_compliance_matrix"}
            yield {"type": "complete", "text": "done", "tools_called": ["query_compliance_matrix"], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="thresholds",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        deltas = [e for e in events if e.get("type") == "tool_input_delta"]
        assert len(deltas) == 0, "compliance matrix deltas should be filtered out"

    @async_test
    async def test_web_search_delta_filtered(self):
        """tool_input_delta for web_search should NOT emit."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            yield {"type": "tool_use", "name": "web_search", "input": {}, "tool_use_id": "tu-5"}
            yield {"type": "tool_input_delta", "tool_use_id": "tu-5", "delta": "q", "name": "web_search"}
            yield {"type": "complete", "text": "", "tools_called": ["web_search"], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="market research",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        deltas = [e for e in events if e.get("type") == "tool_input_delta"]
        assert len(deltas) == 0


class TestToolInputDedup:
    """tool_input without a matching tool_use_id should NOT emit a duplicate card."""

    @async_test
    async def test_tool_input_with_valid_id_emits(self):
        """tool_input with a matching tool_use_id should patch the existing card."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            yield {"type": "tool_use", "name": "search_far", "input": {}, "tool_use_id": "tu-10"}
            yield {"type": "tool_input", "name": "search_far", "input": {"query": "FAR 13"}}
            yield {"type": "tool_result", "name": "search_far", "result": {"text": "done"}}
            yield {"type": "complete", "text": "ok", "tools_called": ["search_far"], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="search",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        tool_uses = [e for e in events if e.get("type") == "tool_use"]
        # Should have 2 tool_use events: initial + patched input
        assert len(tool_uses) == 2
        # Both should carry the same tool_use_id
        ids = [e["tool_use"]["tool_use_id"] for e in tool_uses]
        assert ids[0] == "tu-10"
        assert ids[1] == "tu-10"

    @async_test
    async def test_tool_input_without_id_does_not_emit(self):
        """tool_input when ID queue is exhausted should be silently dropped."""
        from app.streaming_routes import stream_generator

        async def fake_sdk(**kwargs):
            # No preceding tool_use — the ID queue will be empty
            yield {"type": "tool_input", "name": "search_far", "input": {"query": "orphan"}}
            yield {"type": "text", "data": "done"}
            yield {"type": "complete", "text": "done", "tools_called": [], "usage": {}}

        with patch("app.streaming_routes.sdk_query_streaming", fake_sdk), \
             patch("app.streaming_routes.add_message"):
            lines = await _collect_stream(stream_generator(
                message="test",
                tenant_id="t", user_id="u", tier="advanced",
                subscription_service=None, session_id="s",
            ))

        events = _parse_sse_events(lines)
        tool_uses = [e for e in events if e.get("type") == "tool_use"]
        assert len(tool_uses) == 0, "tool_input without matching ID should not create a card"


# ══════════════════════════════════════════════════════════════════════
# PLAN 3 — Statement of Need template generation
# ══════════════════════════════════════════════════════════════════════


class TestSONDocTypeLabels:
    """son_products and son_services must be registered in the generation pipeline."""

    def test_son_products_in_labels(self):
        from app.tools.create_document_support import _DOC_TYPE_LABELS
        assert "son_products" in _DOC_TYPE_LABELS
        assert "Products" in _DOC_TYPE_LABELS["son_products"]

    def test_son_services_in_labels(self):
        from app.tools.create_document_support import _DOC_TYPE_LABELS
        assert "son_services" in _DOC_TYPE_LABELS
        assert "Services" in _DOC_TYPE_LABELS["son_services"]

    def test_son_products_has_system_prompt(self):
        from app.tools.create_document_support import _DOC_TYPE_SYSTEM_PROMPTS
        assert "son_products" in _DOC_TYPE_SYSTEM_PROMPTS
        prompt = _DOC_TYPE_SYSTEM_PROMPTS["son_products"]
        assert "Statement of Need" in prompt
        assert "Products" in prompt
        assert "FAR Part 11" in prompt

    def test_son_services_has_system_prompt(self):
        from app.tools.create_document_support import _DOC_TYPE_SYSTEM_PROMPTS
        assert "son_services" in _DOC_TYPE_SYSTEM_PROMPTS
        prompt = _DOC_TYPE_SYSTEM_PROMPTS["son_services"]
        assert "Statement of Need" in prompt
        assert "Services" in prompt
        assert "FAR Part 11" in prompt

    def test_son_products_prompt_has_required_sections(self):
        from app.tools.create_document_support import _DOC_TYPE_SYSTEM_PROMPTS
        prompt = _DOC_TYPE_SYSTEM_PROMPTS["son_products"]
        for section in [
            "DESCRIPTION OF NEED",
            "PRODUCT SPECIFICATIONS",
            "DELIVERY REQUIREMENTS",
            "ESTIMATED COST",
            "APPROVALS",
        ]:
            assert section in prompt, f"Missing required section: {section}"

    def test_son_services_prompt_has_required_sections(self):
        from app.tools.create_document_support import _DOC_TYPE_SYSTEM_PROMPTS
        prompt = _DOC_TYPE_SYSTEM_PROMPTS["son_services"]
        for section in [
            "DESCRIPTION OF NEED",
            "SERVICE REQUIREMENTS",
            "PERIOD OF PERFORMANCE",
            "PERFORMANCE STANDARDS",
            "APPROVALS",
        ]:
            assert section in prompt, f"Missing required section: {section}"


class TestSONGeneratorFunctions:
    """Generator functions must exist and be wired into the dispatch map."""

    def test_son_products_generator_exists(self):
        from app.tools.create_document_support import _generate_son_products
        assert callable(_generate_son_products)

    def test_son_services_generator_exists(self):
        from app.tools.create_document_support import _generate_son_services
        assert callable(_generate_son_services)

    def test_son_products_in_document_generation_imports(self):
        """The dispatch map in document_generation.py should include son_products."""
        from app.tools.document_generation import exec_create_document
        # Read the source to verify the generator is wired
        import inspect
        source = inspect.getsource(exec_create_document)
        assert "son_products" in source

    def test_son_services_in_document_generation_imports(self):
        from app.tools.document_generation import exec_create_document
        import inspect
        source = inspect.getsource(exec_create_document)
        assert "son_services" in source


class TestSONDocTypeValidation:
    """son_products and son_services should be recognized as valid doc types."""

    def test_son_products_is_valid(self):
        from app.doc_type_registry import is_valid_doc_type
        assert is_valid_doc_type("son_products") is True

    def test_son_services_is_valid(self):
        from app.doc_type_registry import is_valid_doc_type
        assert is_valid_doc_type("son_services") is True

    def test_son_normalizes_from_hyphen(self):
        from app.doc_type_registry import normalize_doc_type
        assert normalize_doc_type("son-products") == "son_products"
        assert normalize_doc_type("son-services") == "son_services"


# ══════════════════════════════════════════════════════════════════════
# PLAN 4 — Package checklist endpoint
# ══════════════════════════════════════════════════════════════════════


class TestPackageChecklistEndpoint:
    """The /api/packages/{id}/checklist endpoint must return document items."""

    def test_checklist_function_exists(self):
        from app.package_store import get_package_checklist
        assert callable(get_package_checklist)

    def test_checklist_route_registered(self):
        """The checklist route must be registered on the packages router."""
        from app.routers.packages import router
        paths = [r.path for r in router.routes]
        assert any("checklist" in p for p in paths)


class TestPackageSessionLinking:
    """Packages must support optional session_id linking."""

    def test_create_package_accepts_session_id(self):
        """create_package function signature must accept session_id."""
        import inspect
        from app.package_store import create_package
        sig = inspect.signature(create_package)
        assert "session_id" in sig.parameters

    def test_session_id_is_optional(self):
        """session_id should default to None."""
        import inspect
        from app.package_store import create_package
        sig = inspect.signature(create_package)
        param = sig.parameters["session_id"]
        assert param.default is None


# ══════════════════════════════════════════════════════════════════════
# PLAN 5 — XLSX document page: markdown above spreadsheet
# ══════════════════════════════════════════════════════════════════════


class TestXlsxMarkdownAboveSpreadsheet:
    """The document viewer XLSX layout must render CollapsibleMarkdown above SpreadsheetPreview."""

    _PAGE_PATH = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..", "client", "app", "documents", "[id]", "page.tsx")
    )

    def _read_page_source(self) -> str:
        with open(self._PAGE_PATH, encoding="utf-8") as f:
            return f.read()

    def test_xlsx_branch_contains_collapsible_markdown(self):
        """The XLSX layout branch must include CollapsibleMarkdown."""
        source = self._read_page_source()
        # Find the XLSX layout comment marker
        xlsx_start = source.find("XLSX Layout:")
        assert xlsx_start != -1, "XLSX Layout comment not found"
        # CollapsibleMarkdown must appear within the XLSX layout block
        xlsx_block = source[xlsx_start:xlsx_start + 2000]
        assert "CollapsibleMarkdown" in xlsx_block, (
            "CollapsibleMarkdown not found in XLSX layout — markdown should render above spreadsheet"
        )

    def test_markdown_renders_before_spreadsheet(self):
        """CollapsibleMarkdown must appear BEFORE SpreadsheetPreview in XLSX layout."""
        source = self._read_page_source()
        xlsx_start = source.find("XLSX Layout:")
        assert xlsx_start != -1
        xlsx_block = source[xlsx_start:xlsx_start + 2000]
        md_pos = xlsx_block.find("CollapsibleMarkdown")
        sp_pos = xlsx_block.find("SpreadsheetPreview")
        assert md_pos < sp_pos, (
            f"CollapsibleMarkdown (pos {md_pos}) must appear before SpreadsheetPreview (pos {sp_pos})"
        )

    def test_markdown_is_conditionally_rendered(self):
        """Markdown block should only render when documentContent exists."""
        source = self._read_page_source()
        xlsx_start = source.find("XLSX Layout:")
        xlsx_block = source[xlsx_start:xlsx_start + 2000]
        assert "documentContent &&" in xlsx_block or "{documentContent &&" in xlsx_block, (
            "CollapsibleMarkdown in XLSX layout must be conditional on documentContent"
        )

    def test_bottom_drawer_chat_still_present(self):
        """Bottom drawer chat must still exist after the spreadsheet."""
        source = self._read_page_source()
        xlsx_start = source.find("XLSX Layout:")
        xlsx_block = source[xlsx_start:xlsx_start + 4000]
        assert "Bottom Drawer" in xlsx_block, "Bottom drawer chat section missing from XLSX layout"


# ══════════════════════════════════════════════════════════════════════
# PLAN 8 — Baseline questions skill updated for 14 questions
# ══════════════════════════════════════════════════════════════════════


class TestBaselineQuestionsSkillUpdate:
    """SKILL.md must document all 14 baseline questions."""

    _SKILL_PATH = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..", ".claude", "skills", "baseline-questions", "SKILL.md")
    )

    def _read_skill(self) -> str:
        with open(self._SKILL_PATH, encoding="utf-8") as f:
            return f.read()

    def test_no_hardcoded_six_questions(self):
        """SKILL.md must not reference '6 questions' as fixed count."""
        content = self._read_skill()
        assert "sends 6 questions" not in content, "Still references '6 questions' as fixed count"
        assert "The 6 Baseline Questions" not in content, "Still has old '6 Baseline Questions' heading"

    def test_no_hardcoded_rows_2_7(self):
        """SKILL.md must not hard-code 'rows 2-7'."""
        content = self._read_skill()
        assert "rows 2-7" not in content, "Still hard-codes 'rows 2-7'"

    def test_documents_14_questions(self):
        """SKILL.md must reference 14 questions."""
        content = self._read_skill()
        assert "14 questions" in content

    def test_q7_sole_source_documented(self):
        """Q7 (Illumina sole-source) must be in the question reference table."""
        content = self._read_skill()
        assert "Q7" in content
        assert "Sole-source" in content or "sole-source" in content

    def test_q13_document_gen_documented(self):
        """Q13 (SOW generation) must be in the question reference table."""
        content = self._read_skill()
        assert "Q13" in content
        assert "Document Gen" in content or "SOW" in content

    def test_q14_cor_gsa_documented(self):
        """Q14 (COR/GSA Purchase) must be in the question reference table."""
        content = self._read_skill()
        assert "Q14" in content
        assert "GSA" in content

    def test_extended_questions_section_exists(self):
        """Must have an Extended Questions section for Q7-Q14."""
        content = self._read_skill()
        assert "Extended Questions" in content

    def test_expected_tool_patterns_include_q7(self):
        """Expected tool patterns must include Q7."""
        content = self._read_skill()
        assert "Q7 (Sole-source" in content or "Q7 (Sole-source J&A)" in content

    def test_expected_tool_patterns_include_q13(self):
        """Expected tool patterns must include Q13."""
        content = self._read_skill()
        assert "Q13 (SOW Generation)" in content

    def test_excel_questions_match_skill_count(self):
        """Excel must have at least 14 questions matching SKILL.md documentation."""
        xlsx_path = os.path.normpath(
            os.path.join(os.path.dirname(__file__), "..", "..", "Use Case List.xlsx")
        )
        if not os.path.exists(xlsx_path):
            pytest.skip("Use Case List.xlsx not found")
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["Baseline questions"]
        count = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=4).value:
                count += 1
        wb.close()
        assert count >= 14, f"Excel has {count} questions, expected at least 14"
